#!/usr/bin/env python3
"""Разовая миграция: существующий xlsx-реестр → SQLite (источник правды).

Читает реестр по пути REGISTRY_FILE и переливает записи в SQLite.
Повторный запуск безопасен: строки сверяются по «ID заявки».

Запуск:  python scripts/migrate_xlsx_to_sqlite.py
         docker compose run --rm app python scripts/migrate_xlsx_to_sqlite.py
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from bot.models import SHEET_HEADERS  # noqa: E402
from config import settings  # noqa: E402
from services import registry_sqlite  # noqa: E402

_ID_COL = SHEET_HEADERS.index("ID заявки")


def migrate() -> tuple[int, int]:
    """Возвращает (импортировано, пропущено)."""
    path = settings.registry_path
    if path is None or not path.exists():
        print(f"Реестр {path} не найден — мигрировать нечего.")
        return 0, 0

    from openpyxl import load_workbook

    ws = load_workbook(path, read_only=True).active
    imported = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = ["" if v is None else str(v) for v in row[: len(SHEET_HEADERS)]]
        if len(values) <= _ID_COL or not values[_ID_COL].startswith("INV-"):
            skipped += 1
            continue
        if registry_sqlite.import_row_sync(values):
            imported += 1
        else:
            skipped += 1
    return imported, skipped


if __name__ == "__main__":
    done, skip = migrate()
    print(f"Импортировано: {done}, пропущено (дубликаты/служебные): {skip}")
