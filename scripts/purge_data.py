#!/usr/bin/env python3
"""Очистка созданных документов и их следов — подготовка к выводу в прод.

По умолчанию НИЧЕГО не удаляет: печатает подробный отчёт «что будет
очищено» — сколько файлов и заявок, размеры, периоды, разбивки по типам
и статусам, примеры записей. Удаление — только с --apply, и перед ним
собирается страховочный бэкап (тот же архив, что /backup), чтобы откат
был возможен.

Данные в docker лежат в /data (см. docker-compose.yml), поэтому запускать
изнутри контейнера:

    docker compose exec app python scripts/purge_data.py                 # отчёт
    docker compose exec app python scripts/purge_data.py --preset all    # отчёт по всему
    docker compose exec app python scripts/purge_data.py --apply --yes   # удалить

Бота на время удаления лучше остановить — он пишет в те же файлы и держит
часть состояния в памяти (настройки, справочник, диалоги):

    docker compose stop app
    docker compose run --rm --no-deps app python scripts/purge_data.py --apply --yes
    docker compose start app

Код возврата: 0 — успех, 1 — очистка не выполнена (нет подтверждения,
google-бэкенд без --allow-google, ошибки удаления).
"""
from __future__ import annotations

import argparse
import json
import os
import pickle
import sqlite3
import sys
from collections import Counter
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import settings  # noqa: E402

# Слово подтверждения: --apply без --yes спрашивает именно его.
CONFIRM_WORD = "ОЧИСТИТЬ"

# Цели пресета по умолчанию: сами документы и следы, которые без них
# бессмысленны (реестр, дедуп, карточки в чатах, причины статусов).
DOCUMENT_KEYS = ("files", "registry", "requests", "reasons", "dedup", "cards")
# Остальное — только по явному выбору: это не документы, а история и
# рабочие справочники, терять их вслепую нельзя.
EXTRA_KEYS = ("audit", "conversations", "incidents", "backups", "users")

# Ошибки удаления — печатаются в конце, не срывают остальные цели.
ERRORS: list[str] = []
# Страховочный архив этого запуска: цель «backups» не должна его съесть.
SAFETY_BACKUP: Path | None = None


# ---------------------------------------------------------------------------
# Мелкие форматтеры
# ---------------------------------------------------------------------------
def _stamp(ts: float | None) -> str:
    if not ts:
        return "—"
    return datetime.fromtimestamp(ts, ZoneInfo(settings.timezone)).strftime("%d.%m.%Y %H:%M")


def _size(num: int) -> str:
    if num < 1024:
        return f"{num} Б"
    if num < 1024 * 1024:
        return f"{num / 1024:.1f} КБ"
    return f"{num / 1024 / 1024:.1f} МБ"


@dataclass
class Scan:
    """Что нашлось по одной цели: счётчик, объём и подробности для отчёта."""

    count: int = 0
    unit: str = "шт"
    size: int = 0
    details: list[str] = field(default_factory=list)
    samples: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class Target:
    """Цель очистки: как посчитать и как удалить."""

    key: str
    title: str
    what: str
    scan: Callable[[], Scan]
    wipe: Callable[[], tuple[int, int]]


# ---------------------------------------------------------------------------
# SQLite (реестр заявок, аудит, дедуп, карточки, причины — один файл)
# ---------------------------------------------------------------------------
def _db_path() -> Path:
    return settings.security_db_path


def _connect() -> sqlite3.Connection | None:
    path = _db_path()
    if not path.exists():
        return None
    return sqlite3.connect(path, timeout=10)


def _tables(conn: sqlite3.Connection) -> set[str]:
    return {
        row[0]
        for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
    }


def _scan_table(table: str, build: Callable[[sqlite3.Connection], Scan]) -> Scan:
    conn = _connect()
    if conn is None:
        return Scan(details=[f"файла БД нет: {_db_path()}"])
    try:
        if table not in _tables(conn):
            return Scan(details=[f"таблицы «{table}» ещё нет — чистить нечего"])
        return build(conn)
    finally:
        conn.close()


def _wipe_table(table: str) -> tuple[int, int]:
    """DELETE + VACUUM. Возвращает (сколько строк, сколько байт освободилось)."""
    conn = _connect()
    if conn is None:
        return 0, 0
    try:
        if table not in _tables(conn):
            return 0, 0
        before = _db_path().stat().st_size
        with conn:
            removed = conn.execute(f'DELETE FROM "{table}"').rowcount
        # VACUUM — вне транзакции: без него файл БД не уменьшается.
        conn.execute("VACUUM")
        return removed, max(0, before - _db_path().stat().st_size)
    except sqlite3.Error as exc:
        ERRORS.append(f"{table}: {exc}")
        return 0, 0
    finally:
        conn.close()


def _range_line(conn: sqlite3.Connection, table: str, column: str = "ts") -> str:
    row = conn.execute(f'SELECT MIN({column}), MAX({column}) FROM "{table}"').fetchone()
    return f"период: {_stamp(row[0])} … {_stamp(row[1])}"


# ---------------------------------------------------------------------------
# Файлы хранилища: счета сотрудников + PDF заявок
# ---------------------------------------------------------------------------
def _registry_files() -> list[Path]:
    """xlsx-реестры (основной и, в google-режиме, зеркало) — отдельная цель."""
    paths = [settings.registry_path]
    mirror = settings.registry_xlsx_path
    if mirror is not None:
        paths.append(mirror)
    seen: set[Path] = set()
    out: list[Path] = []
    for path in paths:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            out.append(path)
    return out


def _storage_files() -> list[Path]:
    root = settings.storage_path
    if not root.exists():
        return []
    skip = {p.resolve() for p in _registry_files()}
    return sorted(p for p in root.rglob("*") if p.is_file() and p.resolve() not in skip)


def _linked_paths() -> set[str]:
    """Значения «Ссылка на счет» из реестра — чтобы показать «сирот»."""
    conn = _connect()
    if conn is None:
        return set()
    try:
        if "requests" not in _tables(conn):
            return set()
        return {str(row[0]) for row in conn.execute("SELECT file_link FROM requests") if row[0]}
    finally:
        conn.close()


def scan_files() -> Scan:
    files = _storage_files()
    scan = Scan(count=len(files), unit="файл(ов)")
    if not files:
        scan.details.append(f"каталог пуст: {settings.storage_path}")
        return scan
    linked = _linked_paths()
    stats = {p: p.stat() for p in files}
    scan.size = sum(st.st_size for st in stats.values())
    reports = [p for p in files if p.name.startswith("INV-")]
    orphans = [p for p in files if str(p) not in linked and str(p.resolve()) not in linked]
    by_ext = Counter(p.suffix.lower().lstrip(".") or "без расширения" for p in files)
    mtimes = [st.st_mtime for st in stats.values()]
    scan.details += [
        f"каталог: {settings.storage_path}",
        f"PDF заявок (INV-…): {len(reports)}",
        f"файлов счетов от сотрудников: {len(files) - len(reports)}",
        "по типам: " + ", ".join(f"{k} — {v}" for k, v in by_ext.most_common()),
        f"период (по времени файла): {_stamp(min(mtimes))} … {_stamp(max(mtimes))}",
        f"не привязано ни к одной заявке реестра: {len(orphans)}",
    ]
    scan.samples = [
        f"{p.name} · {_size(stats[p].st_size)} · {_stamp(stats[p].st_mtime)}"
        for p in sorted(files, key=lambda x: stats[x].st_mtime, reverse=True)
    ]
    return scan


def wipe_files() -> tuple[int, int]:
    removed = freed = 0
    for path in _storage_files():
        try:
            size = path.stat().st_size
            path.unlink()
            removed += 1
            freed += size
        except OSError as exc:
            ERRORS.append(f"{path.name}: {exc}")
    return removed, freed


# ---------------------------------------------------------------------------
# xlsx-реестр
# ---------------------------------------------------------------------------
def _xlsx_rows(path: Path) -> int | None:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        rows = max(0, (wb.active.max_row or 1) - 1)
        wb.close()
        return rows
    except Exception as exc:  # noqa: BLE001 — битый файл тоже подлежит очистке
        ERRORS.append(f"{path.name}: не удалось прочитать ({exc})")
        return None


def scan_registry() -> Scan:
    existing = [p for p in _registry_files() if p.exists()]
    scan = Scan(count=len(existing), unit="файл(ов)")
    if not existing:
        scan.details.append("xlsx-реестра ещё нет — чистить нечего")
        return scan
    for path in existing:
        stat = path.stat()
        scan.size += stat.st_size
        rows = _xlsx_rows(path)
        rows_text = "не прочитался" if rows is None else f"строк без шапки: {rows}"
        scan.details.append(f"{path} · {_size(stat.st_size)} · {rows_text}")
        scan.samples.append(f"{path.name} · изменён {_stamp(stat.st_mtime)}")
    scan.details.append("файл будет удалён; при первой новой заявке создастся заново с шапкой")
    return scan


def wipe_registry() -> tuple[int, int]:
    removed = freed = 0
    for path in _registry_files():
        if not path.exists():
            continue
        try:
            size = path.stat().st_size
            path.unlink()
            removed += 1
            freed += size
        except OSError as exc:
            ERRORS.append(f"{path.name}: {exc}")
    return removed, freed


# ---------------------------------------------------------------------------
# Таблицы SQLite
# ---------------------------------------------------------------------------
def scan_requests() -> Scan:
    def build(conn: sqlite3.Connection) -> Scan:
        rows = conn.execute(
            "SELECT request_id, created_at, status, counterparty, amount, currency, "
            "telegram_id, ts FROM requests ORDER BY ts DESC"
        ).fetchall()
        scan = Scan(count=len(rows), unit="заявк(и)")
        if not rows:
            scan.details.append("реестр пуст")
            return scan
        statuses = Counter(r[2] or "—" for r in rows)
        authors = {r[6] for r in rows if r[6]}
        scan.details += [
            "по статусам: " + ", ".join(f"{k} — {v}" for k, v in statuses.most_common()),
            f"авторов: {len(authors)}",
            _range_line(conn, "requests"),
        ]
        scan.samples = [
            f"{r[0]} · {r[1] or '—'} · {r[2] or '—'} · {r[3] or '—'} · {r[4]} {r[5]}"
            for r in rows
        ]
        return scan

    return _scan_table("requests", build)


def scan_reasons() -> Scan:
    def build(conn: sqlite3.Connection) -> Scan:
        rows = conn.execute(
            "SELECT request_id, status, actor, ts FROM request_reasons ORDER BY ts DESC"
        ).fetchall()
        scan = Scan(count=len(rows), unit="запис(ей)")
        if not rows:
            scan.details.append("причин смены статуса нет")
            return scan
        scan.details.append(_range_line(conn, "request_reasons"))
        scan.samples = [f"{r[0]} · {r[1]} · {r[2]} · {_stamp(r[3])}" for r in rows]
        return scan

    return _scan_table("request_reasons", build)


def scan_dedup() -> Scan:
    def build(conn: sqlite3.Connection) -> Scan:
        rows = conn.execute(
            "SELECT request_id, ts FROM dedup ORDER BY ts DESC"
        ).fetchall()
        scan = Scan(count=len(rows), unit="отпечат(ков)")
        if not rows:
            scan.details.append("отпечатков нет")
            return scan
        scan.details += [
            _range_line(conn, "dedup"),
            f"окно дедупа: {settings.dedup_window_days} дн. — до очистки повтор тестовой "
            "заявки будет считаться дублем",
        ]
        scan.samples = [f"{r[0]} · {_stamp(r[1])}" for r in rows]
        return scan

    return _scan_table("dedup", build)


def scan_cards() -> Scan:
    def build(conn: sqlite3.Connection) -> Scan:
        rows = conn.execute(
            "SELECT request_id, chat_id, message_id FROM finance_cards"
        ).fetchall()
        scan = Scan(count=len(rows), unit="карточ(ек)")
        if not rows:
            scan.details.append("карточек нет")
            return scan
        scan.details += [
            f"заявок: {len({r[0] for r in rows})}, чатов: {len({r[1] for r in rows})}",
        ]
        scan.samples = [f"{r[0]} · чат {r[1]} · сообщение {r[2]}" for r in rows]
        scan.warnings.append(
            "сами сообщения останутся в чатах финансистов: кнопки на них ответят "
            "«Заявка не найдена» — старые карточки лучше удалить в чате руками"
        )
        return scan

    return _scan_table("finance_cards", build)


def scan_audit() -> Scan:
    def build(conn: sqlite3.Connection) -> Scan:
        rows = conn.execute(
            "SELECT event, COUNT(*) FROM audit GROUP BY event ORDER BY 2 DESC"
        ).fetchall()
        total = sum(r[1] for r in rows)
        scan = Scan(count=total, unit="событ(ий)")
        if not total:
            scan.details.append("журнал пуст")
            return scan
        scan.details += [
            "по событиям: " + ", ".join(f"{r[0]} — {r[1]}" for r in rows),
            _range_line(conn, "audit"),
        ]
        scan.samples = [
            f"{_stamp(r[0])} · {r[1]} · {r[2] or '—'} · {r[3] or ''}"
            for r in conn.execute(
                "SELECT ts, event, username, details FROM audit ORDER BY ts DESC"
            ).fetchall()
        ]
        scan.warnings.append(
            "это журнал безопасности: доступы, отказы, удаления. Чистить только "
            "ради «нулевой» истории на старте — восстановить будет нечем"
        )
        return scan

    return _scan_table("audit", build)


# ---------------------------------------------------------------------------
# Файлы данных вне SQLite
# ---------------------------------------------------------------------------
def _conversations_path() -> Path:
    return settings.security_db_path.parent / "conversations.pickle"


def scan_conversations() -> Scan:
    path = _conversations_path()
    if not path.exists():
        return Scan(details=[f"файла нет: {path}"])
    stat = path.stat()
    scan = Scan(count=1, unit="файл", size=stat.st_size)
    try:
        with path.open("rb") as fh:
            data = pickle.load(fh)
    except Exception as exc:  # noqa: BLE001 — битый pickle тоже подлежит очистке
        data = None
        scan.details.append(f"не удалось разобрать: {exc}")
    if isinstance(data, dict):
        convs = data.get("conversations") or {}
        started = sum(len(v) for v in convs.values() if isinstance(v, dict))
        scan.details += [
            f"незакрытых диалогов (черновики форм): {started}",
            f"пользователей с сохранённым состоянием: {len(data.get('user_data') or {})}",
        ]
    scan.details.append(f"файл: {path} · изменён {_stamp(stat.st_mtime)}")
    scan.warnings.append(
        "состояние диалогов бот держит в памяти и сбрасывает в файл периодически — "
        "чистить только на остановленном боте, иначе вернётся"
    )
    return scan


def wipe_conversations() -> tuple[int, int]:
    path = _conversations_path()
    if not path.exists():
        return 0, 0
    try:
        size = path.stat().st_size
        path.unlink()
        return 1, size
    except OSError as exc:
        ERRORS.append(f"{path.name}: {exc}")
        return 0, 0


def scan_incidents() -> Scan:
    path = settings.runtime_settings_path
    if not path.exists():
        return Scan(details=[f"файла настроек нет: {path}"])
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (ValueError, OSError) as exc:
        return Scan(details=[f"не удалось прочитать {path}: {exc}"])
    items = raw.get("incidents") or []
    scan = Scan(count=len(items), unit="запис(ей)")
    if not items:
        scan.details.append("журнал инцидентов пуст")
        return scan
    scan.details += [
        f"с учётом склеенных повторов: {sum(int(x.get('count', 1)) for x in items)}",
        "остальные настройки (финансисты, whitelist, админы, напоминания, бэкап) "
        "не затрагиваются",
    ]
    scan.samples = [
        f"{_stamp(x.get('ts'))} · {x.get('kind', 'other')} · {x.get('title', '')} "
        f"×{int(x.get('count', 1))}"
        for x in sorted(items, key=lambda x: float(x.get("ts", 0.0)), reverse=True)
    ]
    scan.warnings.append(
        "работающий бот держит настройки в памяти и может записать журнал обратно — "
        "чистить на остановленном боте"
    )
    return scan


def wipe_incidents() -> tuple[int, int]:
    from services import runtime_settings as rs

    return rs.clear_incidents(), 0


def _backup_files() -> list[Path]:
    from services import backup

    directory = backup.backup_dir()
    if not directory.exists():
        return []
    safety = SAFETY_BACKUP.resolve() if SAFETY_BACKUP else None
    return sorted(
        p
        for p in directory.glob("invoice-bot-backup-*.tar.gz")
        if p.is_file() and (safety is None or p.resolve() != safety)
    )


def scan_backups() -> Scan:
    files = _backup_files()
    scan = Scan(count=len(files), unit="архив(ов)")
    if not files:
        scan.details.append("архивов нет")
        return scan
    stats = {p: p.stat() for p in files}
    scan.size = sum(st.st_size for st in stats.values())
    scan.details += [
        f"каталог: {files[0].parent}",
        "внутри — копии тестовых заявок и файлов: пока архивы лежат, тестовые "
        "данные из системы не ушли",
        "страховочный бэкап этого запуска не удаляется",
    ]
    scan.samples = [
        f"{p.name} · {_size(stats[p].st_size)} · {_stamp(stats[p].st_mtime)}"
        for p in sorted(files, key=lambda x: stats[x].st_mtime, reverse=True)
    ]
    return scan


def wipe_backups() -> tuple[int, int]:
    removed = freed = 0
    for path in _backup_files():
        try:
            size = path.stat().st_size
            path.unlink()
            removed += 1
            freed += size
        except OSError as exc:
            ERRORS.append(f"{path.name}: {exc}")
    return removed, freed


def scan_users() -> Scan:
    path = settings.user_directory_path
    if not path.exists():
        return Scan(details=[f"справочника нет: {path}"])
    stat = path.stat()
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (ValueError, OSError) as exc:
        return Scan(count=1, unit="файл", size=stat.st_size, details=[f"не читается: {exc}"])
    scan = Scan(count=len(raw), unit="запис(ей)", size=stat.st_size)
    scan.details += [
        f"файл: {path} · изменён {_stamp(stat.st_mtime)}",
        "это карта @username → chat_id: тех, кто в ней есть, бот находит по @username",
    ]
    scan.samples = [f"@{name} → {uid}" for name, uid in sorted(raw.items())]
    scan.warnings.append(
        "после очистки уведомления по @username не дойдут, пока человек снова не "
        "напишет боту; на whitelist и состав финансистов (они по id) не влияет"
    )
    return scan


def wipe_users() -> tuple[int, int]:
    path = settings.user_directory_path
    if not path.exists():
        return 0, 0
    try:
        size = path.stat().st_size
        path.unlink()
        return 1, size
    except OSError as exc:
        ERRORS.append(f"{path.name}: {exc}")
        return 0, 0


# ---------------------------------------------------------------------------
# Реестр целей
# ---------------------------------------------------------------------------
def build_targets() -> list[Target]:
    return [
        Target(
            "files",
            "Файлы документов",
            "PDF заявок и загруженные счета в STORAGE_DIR",
            scan_files,
            wipe_files,
        ),
        Target(
            "registry",
            "xlsx-реестр",
            "файл реестра заявок (зеркало/выгрузка «Экспорт»)",
            scan_registry,
            wipe_registry,
        ),
        Target(
            "requests",
            "Заявки в реестре (SQLite)",
            "источник правды: строки таблицы requests",
            scan_requests,
            lambda: _wipe_table("requests"),
        ),
        Target(
            "reasons",
            "Причины смены статуса",
            "пометки «почему отклонили/отложили»",
            scan_reasons,
            lambda: _wipe_table("request_reasons"),
        ),
        Target(
            "dedup",
            "Отпечатки дедупа",
            "хэши поданных заявок (защита от двойной подачи)",
            scan_dedup,
            lambda: _wipe_table("dedup"),
        ),
        Target(
            "cards",
            "Карточки финансистов",
            "ссылки на сообщения-карточки в чатах",
            scan_cards,
            lambda: _wipe_table("finance_cards"),
        ),
        Target(
            "audit",
            "Аудит-журнал",
            "события доступа, отказов, удалений",
            scan_audit,
            lambda: _wipe_table("audit"),
        ),
        Target(
            "conversations",
            "Состояние диалогов",
            "conversations.pickle: незаконченные черновики форм",
            scan_conversations,
            wipe_conversations,
        ),
        Target(
            "incidents",
            "Журнал инцидентов",
            "список сбоев в карточке «Здоровье бота»",
            scan_incidents,
            wipe_incidents,
        ),
        Target(
            "backups",
            "Архивы бэкапов",
            "tar.gz с копиями тестовых данных",
            scan_backups,
            wipe_backups,
        ),
        Target(
            "users",
            "Справочник пользователей",
            "known_users.json: @username → chat_id",
            scan_users,
            wipe_users,
        ),
    ]


def select_targets(preset: str, only: str, skip: str) -> tuple[list[Target], list[str]]:
    """Возвращает (выбранные цели, неизвестные ключи из --only/--skip)."""
    targets = build_targets()
    known = {t.key for t in targets}
    unknown = sorted(
        k
        for k in {*_split(only), *_split(skip)}
        if k not in known
    )
    if only:
        keys = {k for k in _split(only) if k in known}
    else:
        keys = set(DOCUMENT_KEYS if preset == "documents" else known)
    keys -= set(_split(skip))
    return [t for t in targets if t.key in keys], unknown


def _split(raw: str) -> list[str]:
    return [part.strip() for part in raw.split(",") if part.strip()]


# ---------------------------------------------------------------------------
# Контекст запуска: где данные, кто их сейчас держит
# ---------------------------------------------------------------------------
def _running_bot() -> list[str]:
    """Процессы бота, видимые из этого namespace (в контейнере — main.py, PID 1)."""
    found: list[str] = []
    proc = Path("/proc")
    if not proc.exists():
        return found
    mine = str(os.getpid())
    for entry in proc.iterdir():
        if not entry.name.isdigit() or entry.name == mine:
            continue
        try:
            raw = (entry / "cmdline").read_bytes()
        except OSError:
            continue
        cmd = raw.replace(b"\0", b" ").decode("utf-8", "replace").strip()
        if cmd.endswith("main.py") or " main.py" in cmd:
            found.append(f"pid {entry.name}: {cmd}")
    return found


def context_warnings() -> list[str]:
    """Проблемы окружения, из-за которых отчёт может смотреть не туда."""
    out: list[str] = []
    project_data = Path(__file__).resolve().parent.parent / "data"
    if not settings.storage_path.exists() and (project_data / "storage").exists():
        out.append(
            f"каталог {settings.storage_path} не существует, а {project_data / 'storage'} — есть. "
            "Похоже, скрипт запущен на хосте: в docker данные лежат в /data. "
            "Запускайте через `docker compose exec app python scripts/purge_data.py`"
        )
    if settings.storage_is_google:
        out.append(
            "STORAGE_BACKEND=google: заявки лежат в Google Sheets, файлы — в Drive. "
            "Скрипт чистит только локальные данные; таблицу и папку нужно очищать "
            "вручную. Для --apply требуется флаг --allow-google"
        )
    running = _running_bot()
    if running:
        out.append(
            "бот сейчас работает (" + "; ".join(running) + "). Файлы и БД пишутся "
            "параллельно; для чистого результата остановите его: docker compose stop app"
        )
    return out


def context_lines() -> list[str]:
    return [
        f"бэкенд хранилища: {settings.storage_backend}",
        f"каталог документов: {settings.storage_path}",
        f"xlsx-реестр: {settings.registry_path}",
        f"БД (заявки, аудит, дедуп, карточки): {settings.security_db_path}"
        + (f" · {_size(_db_path().stat().st_size)}" if _db_path().exists() else " · нет файла"),
        f"настройки бота: {settings.runtime_settings_path}",
        f"справочник пользователей: {settings.user_directory_path}",
        f"таймзона: {settings.timezone}",
    ]


# ---------------------------------------------------------------------------
# Отчёт
# ---------------------------------------------------------------------------
def print_report(
    targets: list[Target], scans: dict[str, Scan], limit: int, full: bool
) -> None:
    print("Очистка данных — отчёт (ничего не удалено)\n")
    for line in context_lines():
        print(f"  {line}")
    warnings = context_warnings()
    if warnings:
        print()
        for line in warnings:
            print(f"  ⚠️  {line}")
    print()

    for target in targets:
        scan = scans[target.key]
        head = f"▸ {target.title} [{target.key}] — {scan.count} {scan.unit}"
        if scan.size:
            head += f", {_size(scan.size)}"
        print(head)
        print(f"    что это: {target.what}")
        for line in scan.details:
            print(f"    · {line}")
        shown = scan.samples if full else scan.samples[:limit]
        for line in shown:
            print(f"      – {line}")
        hidden = len(scan.samples) - len(shown)
        if hidden > 0:
            print(f"      … и ещё {hidden} (см. --full или --limit N)")
        for line in scan.warnings:
            print(f"    ⚠️  {line}")
        print()

    total_items = sum(s.count for s in scans.values())
    total_size = sum(s.size for s in scans.values())
    print(f"Итого к удалению: {total_items} объект(ов), освободится ≈ {_size(total_size)}")


def report_json(targets: list[Target], scans: dict[str, Scan], applied: dict) -> str:
    return json.dumps(
        {
            "context": {
                "backend": settings.storage_backend,
                "storage_dir": str(settings.storage_path),
                "registry": str(settings.registry_path),
                "database": str(settings.security_db_path),
                "settings_file": str(settings.runtime_settings_path),
                "user_directory": str(settings.user_directory_path),
                "warnings": context_warnings(),
            },
            "targets": [
                {
                    "key": t.key,
                    "title": t.title,
                    "what": t.what,
                    "count": scans[t.key].count,
                    "unit": scans[t.key].unit,
                    "size_bytes": scans[t.key].size,
                    "details": scans[t.key].details,
                    "samples": scans[t.key].samples,
                    "warnings": scans[t.key].warnings,
                }
                for t in targets
            ],
            "totals": {
                "items": sum(s.count for s in scans.values()),
                "size_bytes": sum(s.size for s in scans.values()),
            },
            "applied": applied,
            "errors": ERRORS,
        },
        ensure_ascii=False,
        indent=2,
    )


# ---------------------------------------------------------------------------
# Удаление
# ---------------------------------------------------------------------------
def confirm(total: int, quiet: bool) -> bool:
    if quiet:
        return True
    if not sys.stdin.isatty():
        print(
            "Нужно подтверждение, а ввод не интерактивный. Добавьте --yes "
            "(например: docker compose exec -T app python scripts/purge_data.py --apply --yes)"
        )
        return False
    typed = input(f"Удалить {total} объект(ов) безвозвратно? Введите {CONFIRM_WORD}: ").strip()
    if typed != CONFIRM_WORD:
        print("Отменено — слово подтверждения не совпало.")
        return False
    return True


def safety_backup() -> Path | None:
    from services import backup

    return backup.create_backup_sync()


def apply_purge(targets: list[Target], quiet: bool, no_backup: bool, total: int) -> dict:
    global SAFETY_BACKUP

    if not confirm(total, quiet):
        return {"done": False, "reason": "нет подтверждения"}

    backup_info: str | None = None
    if no_backup:
        print("⚠️  Страховочный бэкап отключён (--no-backup): откатить будет нечем.\n")
    else:
        try:
            SAFETY_BACKUP = safety_backup()
            backup_info = str(SAFETY_BACKUP)
            print(
                f"Страховочный бэкап: {SAFETY_BACKUP} "
                f"({_size(SAFETY_BACKUP.stat().st_size)})\n"
            )
        except Exception as exc:  # noqa: BLE001 — без бэкапа не удаляем
            print(f"❌ Не удалось собрать бэкап: {exc}")
            print("   Удаление отменено. Осознанно продолжить — --no-backup.")
            return {"done": False, "reason": f"бэкап не собран: {exc}"}

    results: dict[str, dict[str, int]] = {}
    for target in targets:
        removed, freed = target.wipe()
        results[target.key] = {"removed": removed, "freed_bytes": freed}
        print(f"✅ {target.title}: удалено {removed}, освобождено {_size(freed)}")

    return {
        "done": True,
        "backup": backup_info,
        "targets": results,
        "removed": sum(r["removed"] for r in results.values()),
        "freed_bytes": sum(r["freed_bytes"] for r in results.values()),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Отчёт и очистка созданных документов перед выводом в прод.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Цели: " + ", ".join(DOCUMENT_KEYS) + " (пресет documents); "
            + ", ".join(EXTRA_KEYS) + " (только через --preset all или --only)"
        ),
    )
    parser.add_argument(
        "--preset",
        choices=("documents", "all"),
        default="documents",
        help="documents (по умолчанию) — документы и их следы; all — все цели",
    )
    parser.add_argument("--only", default="", help="только эти цели, через запятую")
    parser.add_argument("--skip", default="", help="исключить эти цели, через запятую")
    parser.add_argument("--apply", action="store_true", help="удалить (по умолчанию — только отчёт)")
    parser.add_argument("--yes", action="store_true", help="не спрашивать подтверждение")
    parser.add_argument(
        "--no-backup", action="store_true", help="не собирать страховочный бэкап перед удалением"
    )
    parser.add_argument(
        "--allow-google", action="store_true", help="разрешить --apply при STORAGE_BACKEND=google"
    )
    parser.add_argument("--limit", type=int, default=10, help="сколько примеров показывать (10)")
    parser.add_argument("--full", action="store_true", help="показать все примеры")
    parser.add_argument("--json", action="store_true", help="машинный вывод")
    args = parser.parse_args()

    targets, unknown = select_targets(args.preset, args.only, args.skip)
    if unknown:
        print(f"Неизвестные цели: {', '.join(unknown)}")
        print("Доступны: " + ", ".join(t.key for t in build_targets()))
        return 1
    if not targets:
        print("Не выбрано ни одной цели.")
        return 1

    scans = {t.key: t.scan() for t in targets}
    total = sum(s.count for s in scans.values())

    if not args.apply:
        if args.json:
            print(report_json(targets, scans, {"done": False, "reason": "dry-run"}))
        else:
            print_report(targets, scans, args.limit, args.full)
            print(
                "\nЭто отчёт: ничего не удалено. Удалить — тот же вызов с --apply "
                "(добавьте --yes для неинтерактивного запуска)."
            )
            if ERRORS:
                print("\nПроблемы чтения: " + "; ".join(ERRORS))
        return 0

    if settings.storage_is_google and not args.allow_google:
        print(
            "❌ STORAGE_BACKEND=google: скрипт чистит только локальные данные, "
            "Google Sheets и Drive останутся как есть.\n"
            "   Если это понятно и нужно — повторите с --allow-google."
        )
        return 1

    if not args.json:
        print_report(targets, scans, args.limit, args.full)
        print()
    applied = apply_purge(targets, args.yes, args.no_backup, total)

    if args.json:
        print(report_json(targets, scans, applied))
    elif applied["done"]:
        print(
            f"\nГотово: удалено {applied['removed']} объект(ов), "
            f"освобождено {_size(applied['freed_bytes'])}"
        )
        if ERRORS:
            print("Ошибки: " + "; ".join(ERRORS))

    if not applied["done"]:
        return 1
    return 1 if ERRORS else 0


if __name__ == "__main__":
    sys.exit(main())
