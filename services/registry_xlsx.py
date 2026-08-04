"""Реестр заявок в xlsx (формат шаблона ТЗ «Реестр.xlsx»).

Один модуль обслуживает и ОСНОВНОЙ локальный реестр (STORAGE_BACKEND=local),
и xlsx-зеркало Google-режима — путь к файлу передаётся параметром.
Первые девять колонок листа — ровно по ТЗ; служебные заголовки дописываются
правее только в наш собственный файл (пустой лист).

Вызывать только из services/storage.py: он сериализует доступ блокировкой
и гоняет синхронные функции через to_thread.
"""
from __future__ import annotations

import logging
import os
from collections import Counter
from pathlib import Path

from bot.models import SHEET_HEADERS, InvoiceRequest, excel_safe

log = logging.getLogger(__name__)


def _atomic_save(wb, path: Path) -> None:
    """Сохранение через tmp + os.replace: сбой посреди записи не портит реестр."""
    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    os.replace(tmp, path)

_ID_COL = SHEET_HEADERS.index("ID заявки")
_STATUS_COL = SHEET_HEADERS.index("Статус оплаты")
_COUNTERPARTY_COL = SHEET_HEADERS.index("Контрагент")
_AMOUNT_COL = SHEET_HEADERS.index("Сумма")
_URGENCY_COL = SHEET_HEADERS.index("Срочность")


# ---------------------------------------------------------------------------
# Оформление: реестр должен открываться в Excel как готовый документ.
# ---------------------------------------------------------------------------
def _styles():
    """Ленивая сборка стилей (openpyxl импортируется только при записи)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="D7DEEA")
    return {
        "header_fill": PatternFill("solid", fgColor="1E2A5A"),
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "stripe": PatternFill("solid", fgColor="F4F7FB"),
        "top_left": Alignment(vertical="top", wrap_text=False),
        "wrap": Alignment(vertical="top", wrap_text=True),
        "center": Alignment(horizontal="center", vertical="top"),
        "right": Alignment(horizontal="right", vertical="top"),
        "urgent_font": Font(bold=True, color="C62828"),
        "status_fills": {
            "Оплачена": PatternFill("solid", fgColor="D1E7DD"),
            "Отклонена": PatternFill("solid", fgColor="F8D7DA"),
            "Отложена": PatternFill("solid", fgColor="FFF3CD"),
            "Отозвана": PatternFill("solid", fgColor="E2E3E5"),
        },
    }


# Ширины колонок (по названию — переживают смену порядка).
_COL_WIDTHS = {
    "Дата внесения в реестр": 17,
    "Плановая дата оплаты": 15,
    "Сотрудник по заявке": 26,
    "Контрагент": 28,
    "Сумма": 14,
    "Статья": 22,
    "Статус оплаты": 14,
    "Комментарий": 42,
    "Ссылка на счет": 42,
    "Валюта": 8,
    "Срочность": 11,
    "Реквизиты": 42,
    "ID заявки": 26,
    "Telegram ID": 12,
}

# Колонки с переносом текста и с выравниванием по центру.
_WRAP_TITLES = {"Комментарий", "Реквизиты", "Сотрудник по заявке", "Контрагент", "Статья"}
_CENTER_TITLES = {"Плановая дата оплаты", "Валюта", "Срочность", "Статус оплаты"}


def _style_sheet(ws) -> None:
    """Шапка, ширины, закрепление и автофильтр. Идемпотентно."""
    from openpyxl.utils import get_column_letter

    st = _styles()
    for idx, title in enumerate(SHEET_HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = st["header_fill"]
        cell.font = st["header_font"]
        cell.alignment = st["header_align"]
        cell.border = st["border"]
        width = _COL_WIDTHS.get(title, 16)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SHEET_HEADERS))}{max(ws.max_row, 2)}"
    ws.row_dimensions[1].height = 28


def _style_status_cell(cell) -> None:
    st = _styles()
    fill = st["status_fills"].get(str(cell.value or ""))
    if fill is not None:
        cell.fill = fill


def _style_row(ws, row_idx: int) -> None:
    """Оформление строки данных: границы, зебра, выравнивания, акценты."""
    st = _styles()
    stripe = row_idx % 2 == 0
    for idx, title in enumerate(SHEET_HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=idx)
        cell.border = st["border"]
        if stripe and cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = st["stripe"]
        if title == "Сумма":
            cell.alignment = st["right"]
        elif title in _CENTER_TITLES:
            cell.alignment = st["center"]
        elif title in _WRAP_TITLES:
            cell.alignment = st["wrap"]
        else:
            cell.alignment = st["top_left"]
    urgency_cell = ws.cell(row=row_idx, column=_URGENCY_COL + 1)
    if str(urgency_cell.value or "") == "Срочно":
        urgency_cell.font = st["urgent_font"]
    _style_status_cell(ws.cell(row=row_idx, column=_STATUS_COL + 1))


def _open_or_create(path: Path):
    from openpyxl import Workbook, load_workbook

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист1"
    # Заголовки: пустой лист получает полный набор; лист из шаблона ТЗ
    # (9 колонок) дополняется служебными заголовками правее.
    # ВАЖНО: первую строку проверяем через iter_rows (values_only) — обращение
    # к ws.cell() создаёт ячейку и сдвигает точку вставки ws.append().
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None or all(v in (None, "") for v in first_row):
        for idx, title in enumerate(SHEET_HEADERS, start=1):
            ws.cell(row=1, column=idx).value = title
    else:
        for idx, title in enumerate(SHEET_HEADERS, start=1):
            if ws.cell(row=1, column=idx).value in (None, ""):
                ws.cell(row=1, column=idx).value = title
    return wb, ws


def append_sync(request: InvoiceRequest, path: Path) -> int:
    """Дописывает заявку. Возвращает порядковый номер записи (без шапки)."""
    wb, ws = _open_or_create(path)
    # excel_safe: openpyxl трактует «=...» как формулу — экранируем.
    ws.append([excel_safe(v) for v in request.as_sheet_row()])
    row_idx = ws.max_row
    # Сумма — настоящим числом с денежным форматом (сортировки и суммы в Excel).
    amount_cell = ws.cell(row=row_idx, column=_AMOUNT_COL + 1)
    amount_cell.value = float(request.amount)
    amount_cell.number_format = "#,##0.00"
    _style_sheet(ws)
    _style_row(ws, row_idx)
    _atomic_save(wb, path)
    row_number = row_idx - 1
    log.info("Заявка %s записана в реестр %s (№%s)", request.request_id, path.name, row_number)
    return row_number


def set_status_sync(request_id: str, status_text: str, path: Path) -> dict[str, str] | None:
    """Меняет «Статус оплаты» по ID заявки. Возвращает строку или None."""
    if not path.exists():
        return None
    wb, ws = _open_or_create(path)
    for row in ws.iter_rows(min_row=2):
        cell = row[_ID_COL] if len(row) > _ID_COL else None
        if cell is not None and cell.value == request_id:
            row[_STATUS_COL].value = status_text
            _style_status_cell(row[_STATUS_COL])
            _atomic_save(wb, path)
            log.info("Заявка %s: статус в %s → «%s»", request_id, path.name, status_text)
            values = [
                str(row[i].value) if i < len(row) and row[i].value is not None else ""
                for i in range(len(SHEET_HEADERS))
            ]
            return dict(zip(SHEET_HEADERS, values, strict=True))
    log.warning("Заявка %s не найдена в реестре %s", request_id, path.name)
    return None


def delete_sync(request_id: str, path: Path) -> bool:
    """Удаляет строку заявки из xlsx-зеркала. True — строка была и удалена."""
    if not path.exists():
        return False
    wb, ws = _open_or_create(path)
    for row in ws.iter_rows(min_row=2):
        cell = row[_ID_COL] if len(row) > _ID_COL else None
        if cell is not None and cell.value == request_id:
            ws.delete_rows(cell.row)
            _style_sheet(ws)
            _atomic_save(wb, path)
            log.info("Заявка %s удалена из реестра %s", request_id, path.name)
            return True
    return False


def recent_counterparties_sync(limit: int, path: Path) -> list[str]:
    """Контрагенты из реестра: по частоте, при равенстве — свежее выше."""
    if not path.exists():
        return []
    from openpyxl import load_workbook

    ws = load_workbook(path, read_only=True).active
    counter: Counter[str] = Counter()
    last_pos: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        name = row[_COUNTERPARTY_COL] if len(row) > _COUNTERPARTY_COL else None
        name = str(name).strip() if name else ""
        if name:
            counter[name] += 1
            last_pos[name] = i
    ordered = sorted(counter, key=lambda n: (-counter[n], -last_pos[n]))
    return ordered[:limit]
