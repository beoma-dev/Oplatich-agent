"""Интеграционные тесты HTTP API: полный путь через ASGI, без сети.

Реальные: валидация, дедуп, rate limit, реестр, PDF, фасад хранилища.
Мок — только Telegram-бот (send_message/send_document/get_chat_member).
"""
from __future__ import annotations

import time
from datetime import date, timedelta
from unittest.mock import AsyncMock, MagicMock

import httpx
import pytest

import api.routes as routes_mod
import bot.access as access
import services.runtime_settings as rs
from api.server import build_api
from bot.models import REQUEST_STATUSES
from config import settings
from tests.test_auth import _signed_init_data


def _make_bot() -> MagicMock:
    bot = MagicMock()
    bot.send_message = AsyncMock()
    bot.send_document = AsyncMock()
    member = MagicMock()
    member.status = "member"
    bot.get_chat_member = AsyncMock(return_value=member)
    return bot


@pytest.fixture()
async def api(tmp_paths, monkeypatch):
    # Оба счётчика лимитов — свои на каждый тест: иначе соседние тесты
    # выбирают квоту одного и того же user_id и получают 429.
    monkeypatch.setattr(routes_mod, "_rate", {})
    monkeypatch.setattr(routes_mod, "_my_rate", {})
    monkeypatch.setattr(routes_mod, "_check_rate", {})
    monkeypatch.setattr(access, "_admin_cache", {})
    bot = _make_bot()
    transport = httpx.ASGITransport(app=build_api(bot))
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        yield client, bot


def _auth(user_id: int = 42) -> dict:
    return {
        "X-Telegram-Init-Data": _signed_init_data(
            user={"id": user_id, "first_name": "Тест", "username": "tester"}
        )
    }


def _allow(monkeypatch, ids: str = "42") -> None:
    settings.__dict__.pop("allowed_user_ids", None)
    monkeypatch.setattr(settings, "allowed_user_ids_raw", ids)


def _admins(monkeypatch, ids: str) -> None:
    settings.__dict__.pop("admin_ids", None)
    monkeypatch.setattr(settings, "admin_ids_raw", ids)


def _form(**overrides) -> dict:
    data = {
        "amount": "125 000,50",
        "currency": "RUB",
        "counterparty": "ООО «Ромашка»",
        "article": "Аренда",
        "planned_date": "auto",
        "work_deadline": "текущий месяц",
        "comment": "аренда за июль",
        "urgency": "NORMAL",
        "has_invoice": "0",
        "requisites": "ИНН 7707083893",
    }
    data.update(overrides)
    return data


class TestWorkDeadline:
    """Срок исполнения работ: свободный текст, поле необязательное."""

    async def test_free_text_is_stored(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(work_deadline="поставка в декабре"),
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text
        from openpyxl import load_workbook

        from bot.models import SHEET_HEADERS

        ws = load_workbook(settings.registry_path).active
        # По имени, а не «последняя»: в конец приписываются новые колонки.
        col = SHEET_HEADERS.index("Срок исполнения работ по договору") + 1
        assert ws.cell(2, col).value == "поставка в декабре"

    async def test_absent_field_is_rejected(self, api, monkeypatch):
        """Поле обязательное: без него заявку не принимаем."""
        client, _ = api
        _allow(monkeypatch)
        data = _form()
        data.pop("work_deadline", None)
        resp = await client.post("/api/invoice", data=data, headers=_auth())
        assert resp.status_code == 422

    async def test_blank_field_is_rejected(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice", data=_form(work_deadline="   "), headers=_auth()
        )
        assert resp.status_code == 422

    async def test_too_long_is_rejected(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice", data=_form(work_deadline="я" * 201), headers=_auth()
        )
        # 422 — принятая в этом эндпоинте форма отказа по валидации.
        assert resp.status_code == 422


class TestExtraDocuments:
    """Дополнительные документы: договор, акт, спецификация.

    Необязательны и не зависят от того, есть ли счёт: заявка по реквизитам
    тоже бывает с договором. Формат и размер те же, что у счёта, — разница
    только в количестве.
    """

    def _doc(self, name: str) -> tuple[str, bytes, str]:
        return (name, b"%PDF-1.4 doc", "application/pdf")

    async def test_several_documents_are_saved_and_linked(self, api, monkeypatch):
        from openpyxl import load_workbook

        from bot.models import SHEET_HEADERS

        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(),
            files=[
                ("extra_files", self._doc("dogovor.pdf")),
                ("extra_files", self._doc("akt.pdf")),
            ],
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text

        saved = [p.name for p in settings.storage_path.glob("*доп*")]
        assert len(saved) == 2, saved
        assert any("dogovor" in n for n in saved) and any("akt" in n for n in saved)

        ws = load_workbook(settings.registry_path).active
        col = SHEET_HEADERS.index("Дополнительные документы") + 1
        cell = ws.cell(2, col).value or ""
        assert len(cell.splitlines()) == 2, f"в реестре не обе ссылки: {cell!r}"

    async def test_no_documents_is_normal(self, api, monkeypatch):
        """Подавляющее большинство заявок без них — колонка просто пустая."""
        from openpyxl import load_workbook

        from bot.models import SHEET_HEADERS

        client, _ = api
        _allow(monkeypatch)
        assert (await client.post(
            "/api/invoice", data=_form(), headers=_auth())).status_code == 200
        ws = load_workbook(settings.registry_path).active
        col = SHEET_HEADERS.index("Дополнительные документы") + 1
        assert (ws.cell(2, col).value or "") == ""

    async def test_too_many_is_refused(self, api, monkeypatch):
        from bot.validators import MAX_EXTRA_FILES

        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(),
            files=[("extra_files", self._doc(f"d{i}.pdf")) for i in range(MAX_EXTRA_FILES + 1)],
            headers=_auth(),
        )
        assert resp.status_code == 422
        assert str(MAX_EXTRA_FILES) in resp.json()["detail"]

    async def test_foreign_format_is_refused_by_name(self, api, monkeypatch):
        """В отказе должно быть видно, КАКОЙ файл не подошёл: их несколько."""
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(),
            files=[
                ("extra_files", self._doc("ok.pdf")),
                ("extra_files", ("virus.exe", b"MZ", "application/octet-stream")),
            ],
            headers=_auth(),
        )
        assert resp.status_code == 422
        assert "virus.exe" in resp.json()["detail"]

    async def test_name_cannot_escape_the_storage_dir(self, api, monkeypatch):
        """Имя приходит от пользователя — через него ходят в соседние каталоги."""
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(),
            files=[("extra_files", ("../../evil.pdf", b"%PDF", "application/pdf"))],
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text
        assert not (settings.storage_path.parent.parent / "evil.pdf").exists()
        assert list(settings.storage_path.glob("*доп*")), "файл вообще не сохранился"


class TestHealth:
    async def test_health_ok_after_recent_pulse(self, api):
        from services import health as pulse

        client, _ = api
        pulse.record_ok()
        resp = await client.get("/api/health")
        assert resp.status_code == 200
        assert resp.json()["ok"] is True

    async def test_health_503_when_telegram_stale(self, api, monkeypatch):
        import time

        from services import health as pulse

        client, _ = api
        stale = time.monotonic() - 10_000
        monkeypatch.setattr(pulse, "_last_ok", stale)
        monkeypatch.setattr(pulse, "_started", stale)
        resp = await client.get("/api/health")
        assert resp.status_code == 503
        assert resp.json()["ok"] is False


class TestSubmitAuth:
    async def test_no_init_data_401(self, api):
        client, _ = api
        resp = await client.post("/api/invoice", data=_form())
        assert resp.status_code == 401

    async def test_fail_closed_403(self, api):
        client, _ = api  # whitelist пуст → закрыто
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        assert resp.status_code == 403


class TestSubmitFlow:
    async def test_requisites_flow_writes_registry(self, api, monkeypatch):
        client, bot = api
        _allow(monkeypatch)
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        assert resp.status_code == 200, resp.text
        request_id = resp.json()["request_id"]
        assert request_id.startswith("INV-")
        assert settings.registry_path.exists()
        bot.send_document.assert_awaited()  # подтверждение автору (PDF-подписью)

    async def test_file_flow_saves_invoice(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(has_invoice="1", requisites=""),
            files={"file": ("счёт.pdf", b"%PDF-1.4 test", "application/pdf")},
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text
        saved = list(settings.storage_path.glob("*.pdf"))
        assert any(not p.name.startswith("INV-") for p in saved)  # сам счёт сохранён

    async def test_precheck_endpoint(self, api, monkeypatch):
        from services import invoice_check

        client, _ = api
        _allow(monkeypatch)
        monkeypatch.setattr(
            invoice_check, "inspect_invoice_file", lambda *a, **k: ("⚠️ не счёт", "")
        )
        resp = await client.post(
            "/api/check-file",
            data={"amount": "125 000,50"},
            files={"file": ("cat.jpg", b"\xff\xd8\xff fake", "image/jpeg")},
            headers=_auth(),
        )
        assert resp.status_code == 200
        assert resp.json()["warning"] == "⚠️ не счёт"

        no_auth = await client.post(
            "/api/check-file",
            files={"file": ("cat.jpg", b"\xff\xd8\xff fake", "image/jpeg")},
        )
        assert no_auth.status_code == 401

    async def test_file_warning_returned_to_app(self, api, monkeypatch):
        from services import invoice_check

        client, _ = api
        _allow(monkeypatch)
        monkeypatch.setattr(
            invoice_check, "check_invoice_file", lambda *a, **k: "⚠️ тест: не счёт"
        )
        resp = await client.post(
            "/api/invoice",
            data=_form(has_invoice="1", requisites=""),
            files={"file": ("cat.jpg", b"\xff\xd8\xff fake-jpeg", "image/jpeg")},
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text
        assert resp.json()["warning"] == "⚠️ тест: не счёт"

    async def test_comment_is_optional(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice", data=_form(comment=""), headers=_auth()
        )
        assert resp.status_code == 200, resp.text

    async def test_auto_planned_date_is_server_side(self, api, monkeypatch):
        from openpyxl import load_workbook

        from bot.scheduling import auto_planned_date

        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice", data=_form(urgency="URGENT"), headers=_auth()
        )
        assert resp.status_code == 200
        ws = load_workbook(settings.registry_path).active
        assert ws.cell(2, 2).value == auto_planned_date(True).strftime("%d.%m.%Y")


class TestSubmitValidation:
    @pytest.mark.parametrize(
        "overrides",
        [
            {"amount": "мусор"},
            {"amount": "1.00,50"},
            {"planned_date": "вчера"},
            {"planned_date": "01.01.2020"},
            {"currency": "BTC"},
            {"urgency": "ASAP"},
            {"has_invoice": "2"},
            {"comment": "x" * 501},   # комментарий необязателен, но лимит остался
            {"article": ""},
        ],
    )
    async def test_422(self, api, monkeypatch, overrides):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post("/api/invoice", data=_form(**overrides), headers=_auth())
        assert resp.status_code == 422, resp.text


class TestInvoiceAndRequisitesAreOptional:
    """Ни счёт, ни реквизиты не обязательны (с 26.08.2026).

    Раньше требовалось ровно одно из двух, и заявку «оплатить по договору,
    документы будут позже» подать было нельзя: человек придумывал реквизиты
    или прикладывал что попало, лишь бы форма пропустила. Пустое поле честнее
    выдуманного — финансист видит в карточке, что приложить было нечего.
    """

    async def test_neither_file_nor_requisites(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(has_invoice="0", requisites=""),
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text

    async def test_chose_invoice_but_attached_nothing(self, api, monkeypatch):
        """Выбрал «со счётом» и не приложил — тоже принимаем."""
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(has_invoice="1", requisites=""),
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text

    async def test_card_does_not_promise_a_file_that_is_absent(self, api, monkeypatch):
        """has_invoice в модели — ФАКТ, а не выбор в форме."""
        from openpyxl import load_workbook

        from bot.models import SHEET_HEADERS

        client, bot = api
        _allow(monkeypatch)
        await client.post(
            "/api/invoice", data=_form(has_invoice="1", requisites=""), headers=_auth()
        )
        ws = load_workbook(settings.registry_path).active
        col = SHEET_HEADERS.index("Ссылка на счет") + 1
        assert (ws.cell(2, col).value or "") == "", "ссылка на несуществующий счёт"
        sent = " ".join(
            str(c.kwargs.get("text", "")) + str(c.kwargs.get("caption", ""))
            for c in list(bot.send_message.await_args_list)
            + list(bot.send_document.await_args_list)
        )
        assert "Счёт — этим файлом" not in sent, "карточка обещает вложение, которого нет"

    async def test_requisites_are_still_validated_when_given(self, api, monkeypatch):
        """Необязательное — не значит «любое»: лимит длины остался."""
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice",
            data=_form(has_invoice="0", requisites="я" * 1501),
            headers=_auth(),
        )
        assert resp.status_code == 422


class TestDedupAndRateLimit:
    async def test_duplicate_409_then_force(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        assert (await client.post("/api/invoice", data=_form(), headers=_auth())).status_code == 200
        second = await client.post("/api/invoice", data=_form(), headers=_auth())
        assert second.status_code == 409
        assert second.json()["duplicate"] is True
        forced = await client.post(
            "/api/invoice", data=_form(force="1"), headers=_auth()
        )
        assert forced.status_code == 200

    async def test_rate_limit_429(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        for i in range(5):
            resp = await client.post(
                "/api/invoice",
                data=_form(counterparty=f"Контрагент {i}"),
                headers=_auth(),
            )
            assert resp.status_code == 200
        resp = await client.post(
            "/api/invoice", data=_form(counterparty="шестой"), headers=_auth()
        )
        assert resp.status_code == 429


class TestAdminEndpoints:
    async def test_settings_403_for_non_admin(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.get("/api/admin/settings", headers=_auth())
        assert resp.status_code == 403

    async def test_settings_with_backup_config(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "42")
        resp = await client.get("/api/admin/settings", headers=_auth())
        assert resp.status_code == 200
        body = resp.json()
        assert set(body["backup"]) == {"enabled", "time", "keep", "archive"}
        assert body["registry_url"] is None  # локальный режим — без ссылки

    async def test_archive_contents_are_honest_about_the_backend(self, api, monkeypatch):
        """Панель обязана говорить, чего в архиве НЕТ.

        На google-бэкенде заявки живут в таблице, а счета — в папке Диска,
        и в tar.gz не попадают. Обещание «архив всех данных» человек
        проверил бы в тот единственный день, когда восстанавливается.
        """
        client, _ = api
        _admins(monkeypatch, "42")

        local = (await client.get("/api/admin/settings", headers=_auth())).json()
        assert local["backup"]["archive"]["outside"] == []
        assert any("реестр" in x for x in local["backup"]["archive"]["inside"])

        monkeypatch.setattr(settings, "storage_backend", "google")
        google = (await client.get("/api/admin/settings", headers=_auth())).json()
        outside = google["backup"]["archive"]["outside"]
        assert any("таблиц" in x for x in outside), "не сказано про заявки"
        assert any("Диск" in x for x in outside), "не сказано про счета"
        assert not any("реестр" in x for x in google["backup"]["archive"]["inside"])

    async def test_settings_registry_url_in_google_mode(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "42")
        monkeypatch.setattr(settings, "storage_backend", "google")
        monkeypatch.setattr(settings, "google_sheet_id", "SHEET123")
        resp = await client.get("/api/admin/settings", headers=_auth())
        assert resp.json()["registry_url"].endswith("/SHEET123")

    async def test_settings_drive_url_in_google_mode(self, api, monkeypatch):
        """Рядом с таблицей — папка Диска, куда складываются файлы счетов."""
        client, _ = api
        _admins(monkeypatch, "42")
        monkeypatch.setattr(settings, "storage_backend", "google")
        monkeypatch.setattr(settings, "google_drive_folder_id", "FOLDER123")
        resp = await client.get("/api/admin/settings", headers=_auth())
        assert resp.json()["drive_url"].endswith("/folders/FOLDER123")

    async def test_settings_drive_url_absent_without_folder(self, api, monkeypatch):
        """Папка не задана или хранилище локальное — ссылки нет, кнопка скрыта."""
        client, _ = api
        _admins(monkeypatch, "42")
        assert (await client.get("/api/admin/settings", headers=_auth())
                ).json()["drive_url"] is None
        monkeypatch.setattr(settings, "storage_backend", "google")
        monkeypatch.setattr(settings, "google_drive_folder_id", "")
        assert (await client.get("/api/admin/settings", headers=_auth())
                ).json()["drive_url"] is None

    async def test_backup_save_validation_and_roundtrip(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "42")
        bad = await client.post(
            "/api/admin/backup",
            json={"action": "save", "enabled": True, "time": "25:99", "keep": "7"},
            headers=_auth(),
        )
        assert bad.status_code == 422
        ok = await client.post(
            "/api/admin/backup",
            json={"action": "save", "enabled": False, "time": "04:15", "keep": "3"},
            headers=_auth(),
        )
        assert ok.status_code == 200
        assert ok.json()["backup"] == {"enabled": False, "time": "04:15", "keep": 3}

    async def test_alerts_save_validation_and_roundtrip(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "42")
        bad = await client.post(
            "/api/admin/alerts",
            json={"action": "save", "enabled": True, "link_grace_min": "999"},
            headers=_auth(),
        )
        assert bad.status_code == 422
        ok = await client.post(
            "/api/admin/alerts",
            json={
                "action": "save",
                "enabled": True,
                "kinds": {"backup": False, "storage": False, "выдумка": True},
                "link_grace_min": "12",
            },
            headers=_auth(),
        )
        assert ok.status_code == 200
        cfg = ok.json()["alerts"]
        assert cfg["link_grace_min"] == 12
        assert cfg["kinds"]["backup"] is False
        # Критичное не выключается даже прямым запросом мимо интерфейса.
        assert cfg["kinds"]["storage"] is True
        assert "выдумка" not in cfg["kinds"]
        # Настройки видно там же, где их правят.
        settings_body = (await client.get("/api/admin/settings", headers=_auth())).json()
        assert settings_body["alerts"]["link_grace_min"] == 12
        assert {k["key"] for k in settings_body["alert_kinds"]} == set(rs.ALERT_KEYS)
        assert settings_body["health"]["alive"] in (True, False)

    async def test_alerts_test_message_goes_to_the_asking_admin(self, api, monkeypatch):
        client, bot = api
        _admins(monkeypatch, "42")
        resp = await client.post(
            "/api/admin/alerts", json={"action": "test"}, headers=_auth()
        )
        assert resp.status_code == 200
        assert bot.send_message.call_args.kwargs["chat_id"] == 42

    async def test_alerts_test_reports_closed_chat(self, api, monkeypatch):
        """Бот не может писать первым: честно говорим, что делать."""
        client, bot = api
        _admins(monkeypatch, "42")
        bot.send_message = AsyncMock(side_effect=RuntimeError("bot can't initiate"))
        resp = await client.post(
            "/api/admin/alerts", json={"action": "test"}, headers=_auth()
        )
        assert resp.status_code == 502 and "/start" in resp.json()["detail"]

    async def test_alerts_status_returns_journal(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "42")
        rs.record_incident("backup", "Сбой бэкапа", sent=True, when=time.time())
        body = (await client.post(
            "/api/admin/alerts", json={"action": "status"}, headers=_auth()
        )).json()
        assert body["incidents"][0]["title"] == "Сбой бэкапа"
        assert body["incidents_day"] == 1

    async def test_alerts_are_admin_only(self, api, monkeypatch):
        """Финансисту здоровье бота не показываем: это не его пульт."""
        client, _ = api
        _admins(monkeypatch, "1")
        for body in ({"action": "status"}, {"action": "test"}, {"action": "save"}):
            resp = await client.post("/api/admin/alerts", json=body, headers=_auth(42))
            assert resp.status_code == 403, body

    async def test_env_label_is_empty_on_production(self, api, monkeypatch):
        """Плашка контура — только там, где её задали."""
        client, _ = api
        assert (await client.get("/api/access", headers=_auth(42))).json()["env_label"] == ""
        monkeypatch.setattr(settings, "env_label", "СТЕНД")
        body = (await client.get("/api/access", headers=_auth(42))).json()
        assert body["env_label"] == "СТЕНД"

    async def test_env_label_is_trimmed(self, api, monkeypatch):
        """Метка идёт прямо в разметку — длину ограничиваем на сервере."""
        client, _ = api
        monkeypatch.setattr(settings, "env_label", "  " + "О" * 60 + "  ")
        label = (await client.get("/api/access", headers=_auth(42))).json()["env_label"]
        assert len(label) == 24

    async def test_client_error_reaches_admins(self, api, monkeypatch):
        """Падение формы в браузере больше не остаётся между человеком и им самим."""
        client, bot = api
        _admins(monkeypatch, "1")
        resp = await client.post(
            "/api/client-error",
            json={"message": "TypeError: x is not a function", "where": "app.js:120"},
            headers=_auth(42),
        )
        assert resp.status_code == 200
        text = bot.send_message.call_args.kwargs["text"]
        assert "Ошибка в форме" in text and "app.js:120" in text

    async def test_client_error_is_throttled_per_user(self, api, monkeypatch):
        """Сломанная страница сыплет одним исключением на каждое нажатие."""
        client, bot = api
        _admins(monkeypatch, "1")
        monkeypatch.setattr(routes_mod, "_client_error_seen", {})
        for _ in range(3):
            await client.post(
                "/api/client-error", json={"message": "бум"}, headers=_auth(42)
            )
        assert bot.send_message.call_count == 1
        # ТА ЖЕ ошибка у другого человека — по-прежнему один алерт: одна
        # поломка, а не двадцать. Повторы считает журнал инцидентов.
        await client.post(
            "/api/client-error", json={"message": "бум"}, headers=_auth(43)
        )
        assert bot.send_message.call_count == 1
        # А вот ДРУГАЯ ошибка теряться не должна.
        await client.post(
            "/api/client-error", json={"message": "совсем другое"}, headers=_auth(43)
        )
        assert bot.send_message.call_count == 2

    async def test_client_error_needs_a_message_and_a_signature(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "1")
        empty = await client.post(
            "/api/client-error", json={"message": "  "}, headers=_auth(42)
        )
        assert empty.status_code == 422
        unsigned = await client.post("/api/client-error", json={"message": "бум"})
        assert unsigned.status_code == 401

    async def test_counterparties_after_submit(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        await client.post("/api/invoice", data=_form(), headers=_auth())
        resp = await client.get("/api/counterparties", headers=_auth())
        assert resp.status_code == 200
        names = [it["name"] for it in resp.json()["items"]]
        assert "ООО «Ромашка»" in names


class TestMyRequests:
    """«Мои заявки» через API: выдаются только свои, отзыв меняет статус."""

    async def test_lists_only_own_requests(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch, "42, 43")
        await client.post("/api/invoice", data=_form(), headers=_auth(42))
        await client.post(
            "/api/invoice", data=_form(counterparty="Чужой контрагент"), headers=_auth(43)
        )

        resp = await client.get("/api/my-requests", headers=_auth(42))
        assert resp.status_code == 200
        items = resp.json()["items"]
        assert len(items) == 1
        assert items[0]["counterparty"] == "ООО «Ромашка»"
        assert items[0]["status"] == "Новая"
        # Путь к файлу/ссылка наружу не отдаются — только признак.
        assert "file_url" not in items[0]

    async def test_requires_signed_init_data(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        assert (await client.get("/api/my-requests")).status_code == 401
        assert (
            await client.post("/api/my/withdraw", json={"request_id": "INV-1"})
        ).status_code == 401

    async def test_withdraw_marks_request_and_blocks_repeat(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        request_id = submit.json()["request_id"]

        resp = await client.post(
            "/api/my/withdraw", json={"request_id": request_id}, headers=_auth()
        )
        assert resp.status_code == 200
        assert resp.json()["ok"] is True

        items = (await client.get("/api/my-requests", headers=_auth())).json()["items"]
        assert items[0]["status"] == "Отозвана"

        # Повторный отзыв уже отозванной — отказ с понятным текстом.
        again = await client.post(
            "/api/my/withdraw", json={"request_id": request_id}, headers=_auth()
        )
        assert again.json()["ok"] is False

    async def test_cannot_withdraw_alien_request(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch, "42, 43")
        submit = await client.post("/api/invoice", data=_form(), headers=_auth(43))
        request_id = submit.json()["request_id"]

        resp = await client.post(
            "/api/my/withdraw", json={"request_id": request_id}, headers=_auth(42)
        )
        assert resp.json()["ok"] is False
        assert "только свою" in resp.json()["message"]

    async def test_bad_request_id_rejected(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/my/withdraw", json={"request_id": "'; DROP TABLE requests--"},
            headers=_auth(),
        )
        assert resp.status_code == 422

    async def test_single_request_lookup_for_repeat(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        request_id = submit.json()["request_id"]

        resp = await client.get(
            f"/api/my-requests?request_id={request_id}", headers=_auth()
        )
        items = resp.json()["items"]
        assert len(items) == 1
        # Для повтора важны реквизиты и признак «был ли счёт».
        assert items[0]["requisites"] == "ИНН 7707083893"
        assert items[0]["has_invoice"] is False

    async def test_counterparties_carry_requisites(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        await client.post("/api/invoice", data=_form(), headers=_auth())
        resp = await client.get("/api/counterparties", headers=_auth())
        item = resp.json()["items"][0]
        assert item["name"] == "ООО «Ромашка»"
        assert item["requisites"] == "ИНН 7707083893"


class TestFinancePanel:
    """Панель «Все заявки»: доступ только финансистам, фильтры, итоги."""

    def _financiers(self, monkeypatch, ids: str = "42") -> None:
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", ids)

    async def _submit(self, client, monkeypatch, **overrides):
        resp = await client.post(
            "/api/invoice", data=_form(**overrides), headers=_auth(42)
        )
        # Молчаливый отказ уводил расследование не туда: заявка не создавалась,
        # а падала проверка выборки строк на тридцать ниже.
        assert resp.status_code == 200, f"заявка не создана: {resp.text[:200]}"
        return resp

    async def test_outsider_gets_403(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "999")
        resp = await client.get("/api/finance/requests", headers=_auth(42))
        assert resp.status_code == 403

    async def test_access_flag_tells_ui_to_show_button(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "999")
        assert (await client.get("/api/finance/access", headers=_auth(42))).json() == {
            "ok": False
        }
        self._financiers(monkeypatch, "42")
        assert (await client.get("/api/finance/access", headers=_auth(42))).json() == {
            "ok": True
        }

    async def test_admin_sees_the_panel_without_being_a_financier(self, api, monkeypatch):
        """Список получателей карточек и право видеть заявки — разные вещи.

        Раньше владелец бота, убравший себя из рассылки, заодно терял панель:
        уведомления и доступ решались одним списком. Админ и так видит
        аналитику, аудит и реестр целиком — прятать от него заявки нечем.
        """
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "999")
        _admins(monkeypatch, "42")
        assert (await client.get("/api/finance/requests", headers=_auth(42))).status_code == 200
        assert (await client.get("/api/finance/access", headers=_auth(42))).json() == {
            "ok": True
        }

    async def test_outsider_still_has_no_panel(self, api, monkeypatch):
        """Ни финансист, ни админ — панели нет, отказ пишется в аудит."""
        from services import audit

        client, _ = api
        _allow(monkeypatch, "42,77")
        self._financiers(monkeypatch, "999")
        _admins(monkeypatch, "42")
        assert (await client.get("/api/finance/requests", headers=_auth(77))).status_code == 403
        assert (await client.get("/api/finance/access", headers=_auth(77))).json() == {
            "ok": False
        }
        assert audit.FINANCE_DENIED in [e["event"] for e in await audit.recent_events(5)]

    async def test_admin_may_change_status(self, api, monkeypatch):
        """Кто видит панель, тот и меняет статус: удалять заявку админ и так может."""
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "999")
        _admins(monkeypatch, "42")
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        rid = resp.json()["request_id"]
        out = await client.post("/api/finance/status", headers=_auth(42),
                                json={"request_id": rid, "key": "PAID"})
        assert out.status_code == 200, out.text

    async def test_shows_requests_of_all_authors(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch, "42, 43")
        self._financiers(monkeypatch, "42")
        await self._submit(client, monkeypatch)
        await client.post(
            "/api/invoice", data=_form(counterparty="Чужой контрагент"), headers=_auth(43)
        )

        data = (await client.get("/api/finance/requests", headers=_auth(42))).json()
        names = {it["counterparty"] for it in data["items"]}
        assert names == {"ООО «Ромашка»", "Чужой контрагент"}
        # Финансисту важно, кто подал.
        assert all(it["sender"] for it in data["items"])

    async def test_status_filter(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "42")
        first = await self._submit(client, monkeypatch)
        await self._submit(client, monkeypatch, counterparty="ООО «Вторая»")
        await client.post(
            "/api/my/withdraw",
            json={"request_id": first.json()["request_id"]},
            headers=_auth(42),
        )

        new_only = (await client.get(
            "/api/finance/requests?status=Новая", headers=_auth(42)
        )).json()
        assert new_only["total_found"] == 1
        assert new_only["items"][0]["counterparty"] == "ООО «Вторая»"
        assert new_only["shown"] == 1

        withdrawn = (await client.get(
            "/api/finance/requests?status=Отозвана", headers=_auth(42)
        )).json()
        assert withdrawn["total_found"] == 1

    async def test_query_matches_counterparty_and_article(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "42")
        await self._submit(client, monkeypatch)
        await self._submit(
            client, monkeypatch, counterparty="ИП Петров", article="Хостинг и ПО"
        )

        by_name = (await client.get(
            "/api/finance/requests?query=петров", headers=_auth(42)
        )).json()
        assert by_name["total_found"] == 1
        by_article = (await client.get(
            "/api/finance/requests?query=хостинг", headers=_auth(42)
        )).json()
        assert by_article["total_found"] == 1

    async def test_urgency_filter(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "42")
        await self._submit(client, monkeypatch)
        await self._submit(
            client, monkeypatch, urgency="URGENT", counterparty="ООО «Срочная»"
        )

        urgent = (await client.get(
            "/api/finance/requests?urgency=Срочно", headers=_auth(42)
        )).json()
        assert urgent["total_found"] == 1
        assert urgent["items"][0]["counterparty"] == "ООО «Срочная»"

    async def test_date_range_filter(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "42")
        # Даты считаем от сегодняшнего дня: зашитая «20.08.2026» однажды стала
        # вчерашней, валидатор отверг заявку, и тест начал падать сам по себе.
        target = date.today() + timedelta(days=5)
        await self._submit(
            client, monkeypatch, planned_date=target.strftime("%d.%m.%Y")
        )

        lo = (target - timedelta(days=2)).strftime("%d.%m.%Y")
        hi = (target + timedelta(days=2)).strftime("%d.%m.%Y")
        inside = (await client.get(
            f"/api/finance/requests?date_from={lo}&date_to={hi}",
            headers=_auth(42),
        )).json()
        assert inside["total_found"] == 1

        after = (target + timedelta(days=30)).strftime("%d.%m.%Y")
        outside = (await client.get(
            f"/api/finance/requests?date_from={after}", headers=_auth(42)
        )).json()
        assert outside["total_found"] == 0

    async def test_bad_date_is_rejected(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "42")
        resp = await client.get(
            "/api/finance/requests?date_from=позавчера", headers=_auth(42)
        )
        assert resp.status_code == 422

    async def test_requires_signed_init_data(self, api, monkeypatch):
        client, _ = api
        self._financiers(monkeypatch, "42")
        assert (await client.get("/api/finance/requests")).status_code == 401
        assert (await client.get("/api/finance/access")).status_code == 401


class TestDeleteEndpoint:
    async def test_author_deletes_own_withdrawn_request(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        request_id = submit.json()["request_id"]

        # Пока «Новая» — нельзя.
        first = await client.post(
            "/api/requests/delete", json={"request_id": request_id}, headers=_auth()
        )
        assert first.json()["ok"] is False

        await client.post(
            "/api/my/withdraw", json={"request_id": request_id}, headers=_auth()
        )
        second = await client.post(
            "/api/requests/delete", json={"request_id": request_id}, headers=_auth()
        )
        assert second.json()["ok"] is True
        assert (await client.get("/api/my-requests", headers=_auth())).json()["items"] == []

    async def test_admin_deletes_any_request(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch, "42, 43")
        _admins(monkeypatch, "42")
        submit = await client.post("/api/invoice", data=_form(), headers=_auth(43))
        request_id = submit.json()["request_id"]

        resp = await client.post(
            "/api/requests/delete", json={"request_id": request_id}, headers=_auth(42)
        )
        assert resp.json()["ok"] is True

    async def test_stranger_cannot_delete(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch, "42, 43")
        submit = await client.post("/api/invoice", data=_form(), headers=_auth(43))
        request_id = submit.json()["request_id"]

        resp = await client.post(
            "/api/requests/delete", json={"request_id": request_id}, headers=_auth(42)
        )
        assert resp.json()["ok"] is False

    async def test_bad_id_and_unsigned(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        bad = await client.post(
            "/api/requests/delete", json={"request_id": "DROP TABLE"}, headers=_auth()
        )
        assert bad.status_code == 422
        assert (
            await client.post("/api/requests/delete", json={"request_id": "INV-1"})
        ).status_code == 401


class TestOverdueFlag:
    """Просрочка должна быть видна в списке, а не только в фильтре.

    Признак существовал лишь внутри фильтра «Просрочены»: в списке заявка
    со вчерашним сроком выглядела обычной «Новой», и узнать о сроке можно
    было, только заранее заподозрив и переключив фильтр. Считаем на сервере
    — PENDING_STATUSES и разбор даты живут там, копия логики в JS стала бы
    четвёртым местом, которое разъедется.
    """

    def _row(self, planned: str, status: str = "Новая") -> dict:
        return {
            "ID заявки": "INV-1", "Статус оплаты": status,
            "Плановая дата оплаты": planned, "Сумма": "100.00",
        }

    def test_yesterday_is_overdue(self, monkeypatch):
        from api import routes

        yesterday = (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
        assert routes._is_overdue(self._row(yesterday)) is True
        assert routes._as_item(self._row(yesterday), "")["overdue"] is True

    def test_today_is_not_overdue_yet(self):
        from api import routes

        today = date.today().strftime("%d.%m.%Y")
        assert routes._is_overdue(self._row(today)) is False

    def test_paid_is_never_overdue(self):
        """Оплаченную вчерашним сроком дёргать незачем."""
        from api import routes

        yesterday = (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
        assert routes._is_overdue(self._row(yesterday, "Оплачена")) is False

    def test_missing_date_is_not_overdue(self):
        from api import routes

        assert routes._is_overdue(self._row("")) is False

    def test_filter_and_flag_agree(self):
        """Фильтр «Просрочены» и признак в карточке — одна логика."""
        from api import routes

        yesterday = (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
        row = self._row(yesterday)
        shown = routes._matches(
            row, status=routes.OVERDUE_FILTER, urgency="", query="",
            date_from=None, date_to=None,
        )
        assert shown is routes._is_overdue(row) is True


class TestStaticPage:
    async def test_form_page_must_be_revalidated(self, api):
        """Страница обязана проверяться на сервере при каждом открытии.

        WebView Telegram держит её цепко: без указания пользователь неделями
        видел бы старый JS после деплоя. Раньше здесь стоял no-store, но он
        строже, чем нужно: no-cache требует спросить сервер так же, а
        неизменный файл возвращается как 304 вместо 82 КБ тела (reports/005,
        R20). Ослаблять до must-revalidate с max-age нельзя — вот граница.
        """
        client, _ = api
        resp = await client.get("/")
        assert resp.status_code == 200
        assert resp.headers["cache-control"] == "no-cache"
        assert "<title>" in resp.text or "id=\"form-view\"" in resp.text

    async def test_api_answers_are_never_stored(self, api):
        """А вот ответы API кешировать нельзя вовсе: в них данные заявок."""
        client, _ = api
        resp = await client.get("/api/health")
        assert resp.headers["cache-control"] == "no-store"


class TestFinanceStatusEndpoint:
    """Смена статуса из панели — тот же сценарий, что кнопки в чате."""

    def _financiers(self, monkeypatch, ids: str = "42") -> None:
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", ids)

    async def test_financier_marks_paid(self, api, monkeypatch):
        client, bot = api
        _allow(monkeypatch)
        self._financiers(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        request_id = submit.json()["request_id"]

        resp = await client.post(
            "/api/finance/status",
            json={"request_id": request_id, "key": "PAID"},
            headers=_auth(),
        )
        assert resp.status_code == 200
        assert resp.json()["ok"] is True

        items = (await client.get("/api/my-requests", headers=_auth())).json()["items"]
        assert items[0]["status"] == "Оплачена"
        # Автор узнаёт о смене статуса.
        assert any(
            "Статус вашей заявки обновлён" in (c.kwargs.get("text") or "")
            for c in bot.send_message.await_args_list
        )

    async def test_reason_reaches_the_author_and_the_list(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        request_id = submit.json()["request_id"]

        await client.post(
            "/api/finance/status",
            json={"request_id": request_id, "key": "REJECTED", "reason": "нет бюджета"},
            headers=_auth(),
        )
        items = (await client.get("/api/my-requests", headers=_auth())).json()["items"]
        assert items[0]["status"] == "Отклонена"
        assert items[0]["reason"] == "нет бюджета"

    async def test_outsider_cannot_change_status(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch, "999")
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        resp = await client.post(
            "/api/finance/status",
            json={"request_id": submit.json()["request_id"], "key": "PAID"},
            headers=_auth(),
        )
        assert resp.status_code == 403

    async def test_withdrawn_request_is_locked(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        request_id = submit.json()["request_id"]
        await client.post(
            "/api/my/withdraw", json={"request_id": request_id}, headers=_auth()
        )

        resp = await client.post(
            "/api/finance/status",
            json={"request_id": request_id, "key": "PAID"},
            headers=_auth(),
        )
        assert resp.json()["ok"] is False
        assert "отозвана" in resp.json()["message"].lower()

    async def test_bad_key_and_unsigned(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._financiers(monkeypatch)
        submit = await client.post("/api/invoice", data=_form(), headers=_auth())
        bad = await client.post(
            "/api/finance/status",
            json={"request_id": submit.json()["request_id"], "key": "APPROVED"},
            headers=_auth(),
        )
        assert bad.status_code == 422
        assert (
            await client.post("/api/finance/status", json={"request_id": "INV-1"})
        ).status_code == 401


class TestReminderSettings:
    """Общих настроек напоминаний больше нет: расписание у каждого своё.

    Ручка админа удалена вместе с карточкой — иначе осталась бы точка,
    которая молча меняет расписание всем сразу.
    """

    async def test_shared_endpoint_is_gone(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "42")
        resp = await client.post(
            "/api/admin/reminders", json={"action": "run"}, headers=_auth()
        )
        # 404 или 405 — важно, что ручка больше ничего не делает.
        assert resp.status_code in (404, 405), resp.status_code


class TestAutofillBeta:
    """Бета: разбор счёта отдаётся форме как ПРЕДЛОЖЕНИЕ, не как действие."""

    INVOICE = (
        "Счет на оплату № 118 от 04.08.2026\n"
        "Поставщик (Исполнитель): ООО «Ромашка»\n"
        "ИНН 7707083893  КПП 773601001\n"
        "БИК 044525225\nР/с 40702810400000012345\n"
        "Всего к оплате: 174387.21\n"
    )

    def _text(self, monkeypatch, text: str = "") -> None:
        from services import invoice_check

        monkeypatch.setattr(
            invoice_check, "_extract_text", lambda c, f: (text or self.INVOICE, True)
        )

    async def test_precheck_returns_parsed_fields(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        self._text(monkeypatch)

        resp = await client.post(
            "/api/check-file",
            files={"file": ("s.pdf", b"%PDF-1.4", "application/pdf")},
            headers=_auth(),
        )
        assert resp.status_code == 200
        data = resp.json()["autofill"]
        assert data["amount"] == "174387.21"
        assert data["counterparty"] == "ООО «Ромашка»"
        assert data["inn"] == "7707083893"
        assert "Р/с 40702810400000012345" in data["requisites"]

    async def test_switched_off_returns_nothing(self, api, monkeypatch):
        """Выключили бету — форма ведёт себя ровно как раньше."""
        from services import runtime_settings as rs

        client, _ = api
        _allow(monkeypatch)
        self._text(monkeypatch)
        rs.set_autofill(False)

        resp = await client.post(
            "/api/check-file",
            files={"file": ("s.pdf", b"%PDF-1.4", "application/pdf")},
            headers=_auth(),
        )
        assert resp.json()["autofill"] == {}
        assert "warning" in resp.json()      # основная проверка не пострадала

    async def test_unreadable_file_gives_empty_autofill(self, api, monkeypatch):
        from services import invoice_check

        client, _ = api
        _allow(monkeypatch)
        monkeypatch.setattr(invoice_check, "_extract_text", lambda c, f: ("", False))

        resp = await client.post(
            "/api/check-file",
            files={"file": ("cat.jpg", b"\xff\xd8\xff", "image/jpeg")},
            headers=_auth(),
        )
        assert resp.status_code == 200
        assert resp.json()["autofill"] == {}

    async def test_toggle_requires_admin(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "999")
        assert (await client.post(
            "/api/admin/autofill", json={"enabled": False}, headers=_auth()
        )).status_code == 403

    async def test_admin_toggles_and_settings_report_it(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "42")

        off = await client.post(
            "/api/admin/autofill", json={"enabled": False}, headers=_auth()
        )
        assert off.json()["autofill"] is False
        settings_now = await client.get("/api/admin/settings", headers=_auth())
        assert settings_now.json()["autofill"] is False

        on = await client.post(
            "/api/admin/autofill", json={"enabled": True}, headers=_auth()
        )
        assert on.json()["autofill"] is True


class TestUsersRoster:
    """Список «кто пользуется ботом» и управление админами."""

    async def test_roster_marks_roles_and_access(self, api, monkeypatch):
        from services import runtime_settings as rs

        client, _ = api
        _admins(monkeypatch, "42")
        settings.__dict__.pop("allowed_user_ids", None)
        monkeypatch.setattr(settings, "allowed_user_ids_raw", "7")
        rs.add_allowed(8)
        resp = await client.get("/api/admin/users", headers=_auth())
        assert resp.status_code == 200
        body = resp.json()
        by_id = {u["id"]: u for u in body["users"]}
        assert by_id[7]["access"] == "env"        # запись из .env, но отозвать можно
        assert by_id[8]["access"] == "dynamic"    # добавлена из панели
        assert by_id[42]["admin"] is True
        assert by_id[42]["admin_source"] == "env"
        assert by_id[42]["access"] is None        # админ подаёт и без whitelist
        assert body["whitelist_empty"] is False

    async def test_roster_hides_the_bot_and_chats(self, api, monkeypatch):
        """В списке только люди: ни самого бота, ни групп-получателей."""
        from services import runtime_settings as rs
        from services import user_directory

        client, _ = api
        _admins(monkeypatch, "42")
        user_directory._cache = {"self": settings.bot_id, "vasya": 55}
        rs.add_financier("-1001234567890")     # группа финансистов, не человек
        ids = {u["id"] for u in
               (await client.get("/api/admin/users", headers=_auth())).json()["users"]}
        assert 55 in ids
        assert settings.bot_id not in ids
        assert -1001234567890 not in ids

    async def test_roster_reports_empty_whitelist(self, api, monkeypatch):
        """Пустой whitelist — подача закрыта всем, кроме админов."""
        client, _ = api
        _admins(monkeypatch, "42")
        settings.__dict__.pop("allowed_user_ids", None)
        monkeypatch.setattr(settings, "allowed_user_ids_raw", "")
        body = (await client.get("/api/admin/users", headers=_auth())).json()
        assert body["whitelist_empty"] is True
        assert all(u["access"] is None for u in body["users"])

    async def test_roster_requires_admin(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.get("/api/admin/users", headers=_auth())
        assert resp.status_code == 403

    async def test_env_access_can_be_revoked(self, api, monkeypatch):
        """Отзыв записи из .env: человеку снова придётся просить доступ."""
        from bot.access import is_allowed

        client, _ = api
        _admins(monkeypatch, "42")
        settings.__dict__.pop("allowed_user_ids", None)
        monkeypatch.setattr(settings, "allowed_user_ids_raw", "7")
        assert is_allowed(7)
        resp = await client.post(
            "/api/admin/allowed", json={"action": "remove", "entry": "7"}, headers=_auth()
        )
        assert resp.json()["ok"] is True
        assert not is_allowed(7)
        # Возвращается обычным «добавить».
        await client.post(
            "/api/admin/allowed", json={"action": "add", "entry": "7"}, headers=_auth()
        )
        assert is_allowed(7)

    async def test_admin_can_be_appointed_and_removed(self, api, monkeypatch):
        from bot.access import is_admin

        client, _ = api
        _admins(monkeypatch, "42")
        add = await client.post(
            "/api/admin/admins", json={"action": "add", "entry": "77"}, headers=_auth()
        )
        assert add.json()["ok"] is True
        assert is_admin(77)
        drop = await client.post(
            "/api/admin/admins", json={"action": "remove", "entry": "77"}, headers=_auth()
        )
        assert drop.json()["ok"] is True
        assert not is_admin(77)

    async def test_env_admin_cannot_be_demoted(self, api, monkeypatch):
        """Владелец из .env остаётся владельцем — даже для другого админа."""
        from bot.access import is_admin

        client, _ = api
        _admins(monkeypatch, "42")
        await client.post(
            "/api/admin/admins", json={"action": "add", "entry": "77"}, headers=_auth()
        )
        resp = await client.post(
            "/api/admin/admins", json={"action": "remove", "entry": "42"}, headers=_auth()
        )
        assert resp.status_code == 409
        assert ".env" in resp.json()["detail"]
        assert is_admin(42)

    async def test_admins_endpoint_requires_admin(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/admin/admins", json={"action": "add", "entry": "1"}, headers=_auth()
        )
        assert resp.status_code == 403


class TestAccessRequests:
    """Запрос доступа сотрудником и решение админа."""

    async def test_state_reports_denied_and_pending(self, api, monkeypatch):
        client, _ = api
        _admins(monkeypatch, "1")
        first = (await client.get("/api/access", headers=_auth())).json()
        assert first == {
            "allowed": False, "financier": False, "admin": False,
            "pending": False, "has_admins": True, "env_label": "",
            # Плашка «технические работы» едет тем же ответом: её видят все,
            # кто открыл форму, и ради неё не стоит второго запроса.
            "maintenance": {"enabled": False, "text": rs.MAINTENANCE_DEFAULT},
            # Флаг обкатки живой инструкции: в бою выключен.
            "animated_help": False,
        }
        await client.post("/api/access/request", headers=_auth())
        assert (await client.get("/api/access", headers=_auth())).json()["pending"] is True

    async def test_request_notifies_every_admin_once(self, api, monkeypatch):
        client, bot = api
        _admins(monkeypatch, "1,2")
        resp = await client.post("/api/access/request", headers=_auth())
        assert resp.status_code == 200
        assert bot.send_message.await_count == 2
        # Повторное нажатие админов больше не беспокоит.
        again = await client.post("/api/access/request", headers=_auth())
        assert "уже отправлена" in again.json()["message"]
        assert bot.send_message.await_count == 2

    async def test_request_without_admins_says_so(self, api, monkeypatch):
        client, bot = api
        _admins(monkeypatch, "")
        resp = await client.post("/api/access/request", headers=_auth())
        assert "не задан ни один админ" in resp.json()["message"]
        bot.send_message.assert_not_awaited()

    async def test_request_needs_signature(self, api):
        client, _ = api
        assert (await client.post("/api/access/request")).status_code == 401
        assert (await client.get("/api/access")).status_code == 401

    async def test_allowed_user_gets_no_request(self, api, monkeypatch):
        client, bot = api
        _allow(monkeypatch)
        resp = await client.post("/api/access/request", headers=_auth())
        assert resp.json()["ok"] is False
        bot.send_message.assert_not_awaited()

    async def test_approval_opens_access_and_answers_the_author(self, api, monkeypatch):
        from bot.access import is_allowed
        from services.access_requests import resolve_access

        client, bot = api
        _admins(monkeypatch, "1")
        await client.post("/api/access/request", headers=_auth())
        assert not is_allowed(42)
        note = await resolve_access(bot, 42, True, actor_id=1, actor_name="@boss")
        assert "Доступ открыт" in note
        assert is_allowed(42)
        # Заявка снята — повторная просьба снова дойдёт до админов.
        assert (await client.get("/api/access", headers=_auth())).json()["pending"] is False
        assert any("Доступ открыт" in str(c) for c in bot.send_message.await_args_list)

    async def test_rejection_leaves_access_closed(self, api, monkeypatch):
        from bot.access import is_allowed
        from services.access_requests import resolve_access

        client, bot = api
        _admins(monkeypatch, "1")
        await client.post("/api/access/request", headers=_auth())
        await resolve_access(bot, 42, False, actor_id=1, actor_name="@boss")
        assert not is_allowed(42)
        assert (await client.get("/api/access", headers=_auth())).json()["pending"] is False

    async def test_request_remembers_the_username_for_the_whitelist(self, api, monkeypatch):
        """После выдачи доступа в настройках виден @ник, а не «id 42».

        Whitelist хранит числовые id (ник можно сменить), а показывает
        справочник; из Mini App апдейта в чат нет, поэтому ник туда попадает
        только отсюда.
        """
        from services.access_requests import resolve_access
        from services.user_directory import username_for

        client, bot = api
        _admins(monkeypatch, "1")
        assert username_for(42) is None
        await client.post("/api/access/request", headers=_auth())
        assert username_for(42) == "@tester"

        note = await resolve_access(bot, 42, True, actor_id=1, actor_name="@boss")
        assert "@tester" in note, "решение админа тоже должно называть человека по нику"
        body = (await client.get("/api/admin/settings", headers=_auth(1))).json()
        rows = [u for u in body["allowed"] if u["id"] == 42]
        assert rows and rows[0]["username"] == "@tester"


class TestPersonalReminders:
    """Каждый получатель настраивает напоминания себе, а не всем сразу."""

    async def test_only_recipients_may_read_and_save(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)          # обычный сотрудник: напоминать нечего
        assert (await client.get("/api/reminders/me", headers=_auth())).status_code == 403
        assert (await client.post("/api/reminders/me", json={},
                                  headers=_auth())).status_code == 403

    async def test_financier_sets_own_schedule(self, api, monkeypatch):
        from services import runtime_settings as rs

        client, _ = api
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "42")

        first = (await client.get("/api/reminders/me", headers=_auth())).json()
        assert first["custom"] is False, "пока ничего не меняли — общие настройки"

        saved = await client.post("/api/reminders/me", json={
            "enabled": True, "time": "07:15", "days_before": "3",
            "overdue_enabled": False,
        }, headers=_auth())
        assert saved.status_code == 200
        cfg = saved.json()["reminders"]
        assert (cfg["time"], cfg["days_before"], cfg["overdue_enabled"]) == ("07:15", 3, False)
        assert cfg["custom"] is True
        # Общие настройки при этом не тронуты — это и есть смысл личных.
        assert rs.reminders_config()["time"] != "07:15"

        back = await client.post("/api/reminders/me", json={"action": "reset"},
                                 headers=_auth())
        assert back.json()["reminders"]["custom"] is False

    async def test_personal_time_is_validated(self, api, monkeypatch):
        client, _ = api
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "42")
        bad_time = await client.post("/api/reminders/me", json={"time": "25:99"},
                                     headers=_auth())
        assert bad_time.status_code == 422
        bad_days = await client.post("/api/reminders/me", json={"days_before": "99"},
                                     headers=_auth())
        assert bad_days.status_code == 422

class TestHelpCommand:
    """/help показывает только то, что доступно этому человеку."""

    async def _help(self, bot, user_id: int) -> str:
        from bot.commands import build_help

        return await build_help(bot, user_id)

    async def test_plain_employee_sees_common_commands_only(self, api, monkeypatch):
        _, bot = api
        _allow(monkeypatch)
        text = await self._help(bot, 42)
        assert "/invoice" in text and "/my" in text and "/myid" in text
        assert "/allow" not in text, "админские команды не для всех"
        assert "Финансисту" not in text

    async def test_financier_sees_their_block(self, api, monkeypatch):
        _, bot = api
        _allow(monkeypatch)
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "42")
        text = await self._help(bot, 42)
        assert "Финансисту" in text
        assert "/allow" not in text

    async def test_admin_sees_everything(self, api, monkeypatch):
        _, bot = api
        _admins(monkeypatch, "42")
        text = await self._help(bot, 42)
        assert "Админу" in text
        for command in ("/allow", "/deny", "/fin_add", "/export", "/audit", "/backup"):
            assert command in text, command

    async def test_without_access_help_says_how_to_get_it(self, api):
        _, bot = api
        text = await self._help(bot, 42)     # whitelist пуст → доступа нет
        assert "Запросить доступ" in text


class TestOverdueFilter:
    """Фильтр «Просрочены»: срок прошёл, а заявка всё ещё ждёт оплаты."""

    async def _seed(self, request_id: str, days: int):
        from datetime import date, timedelta

        from services import storage
        from tests.conftest import make_request

        await storage.append_invoice(make_request(
            request_id=request_id, planned_date=date.today() + timedelta(days=days)))

    async def test_overdue_is_a_state_not_a_status(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "42")
        await self._seed("INV-20260804-000001-0001", -3)   # срок прошёл
        await self._seed("INV-20260804-000002-0002", +5)   # ещё впереди

        resp = await client.get("/api/finance/requests?status=__overdue__", headers=_auth())
        assert resp.status_code == 200
        assert [r["id"] for r in resp.json()["items"]] == ["INV-20260804-000001-0001"]

    async def test_paid_request_is_never_overdue(self, api, monkeypatch):
        from services import storage

        client, _ = api
        _allow(monkeypatch)
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "42")
        await self._seed("INV-20260804-000003-0003", -3)
        await storage.set_request_status("INV-20260804-000003-0003", "Оплачена")
        resp = await client.get("/api/finance/requests?status=__overdue__", headers=_auth())
        assert resp.json()["items"] == []


class TestMaintenanceBanner:
    """Плашка «технические работы»: вешает админ, видят все.

    Подачу НЕ блокирует: заявка всё равно уходит в реестр. Молча не принять
    заполненную форму хуже, чем принять её во время работ.
    """

    async def test_off_by_default(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        state = (await client.get("/api/access", headers=_auth())).json()
        assert state["maintenance"]["enabled"] is False

    async def test_admin_turns_it_on_and_everyone_sees_it(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "42")
        resp = await client.post(
            "/api/admin/maintenance",
            json={"enabled": True, "text": "  Обновляем   реестр  "},
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text
        # Пробелы схлопываются: текст идёт в плашку как есть.
        assert resp.json()["maintenance"]["text"] == "Обновляем реестр"

        seen = (await client.get("/api/access", headers=_auth(99))).json()
        assert seen["maintenance"] == {"enabled": True, "text": "Обновляем реестр"}

    async def test_submission_still_works_during_maintenance(self, api, monkeypatch):
        """Плашка предупреждает, а не запрещает — иначе теряется работа."""
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "42")
        await client.post(
            "/api/admin/maintenance", json={"enabled": True}, headers=_auth()
        )
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        assert resp.status_code == 200, resp.text

    async def test_only_admin_may_switch_it(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "1")
        resp = await client.post(
            "/api/admin/maintenance", json={"enabled": True}, headers=_auth()
        )
        assert resp.status_code == 403
        assert (await client.get("/api/access", headers=_auth())).json()[
            "maintenance"
        ]["enabled"] is False

    async def test_empty_text_falls_back_to_the_default_wording(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "42")
        resp = await client.post(
            "/api/admin/maintenance", json={"enabled": True, "text": "   "},
            headers=_auth(),
        )
        assert resp.json()["maintenance"]["text"] == rs.MAINTENANCE_DEFAULT


class TestClosingDocuments:
    """Закрывающие документы: акт, УПД, накладная.

    Приходят ПОСЛЕ оплаты, иногда через месяц, и дописываются в уже
    существующую строку реестра — а не создают новую заявку.
    """

    def _doc(self, name: str) -> tuple[str, bytes, str]:
        return (name, b"%PDF-1.4 akt", "application/pdf")

    async def _submit(self, client, monkeypatch) -> str:
        _allow(monkeypatch)
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        assert resp.status_code == 200, resp.text
        return resp.json()["request_id"]

    async def test_author_attaches_and_registry_row_is_updated(self, api, monkeypatch):
        from openpyxl import load_workbook

        from bot.models import SHEET_HEADERS

        client, _ = api
        rid = await self._submit(client, monkeypatch)
        resp = await client.post(
            "/api/my/closing-docs",
            data={"request_id": rid},
            files=[("files", self._doc("akt.pdf"))],
            headers=_auth(),
        )
        assert resp.status_code == 200, resp.text
        assert resp.json()["count"] == 1

        ws = load_workbook(settings.registry_path).active
        col = SHEET_HEADERS.index("Закрывающие документы") + 1
        assert (ws.cell(2, col).value or "").strip(), "строка реестра не обновилась"

    async def test_second_upload_appends_and_does_not_replace(self, api, monkeypatch):
        """Документы носят частями: вторая загрузка не стирает первую."""
        client, _ = api
        rid = await self._submit(client, monkeypatch)
        for name in ("akt.pdf", "upd.pdf"):
            resp = await client.post(
                "/api/my/closing-docs",
                data={"request_id": rid},
                files=[("files", self._doc(name))],
                headers=_auth(),
            )
            assert resp.status_code == 200, resp.text
        assert resp.json()["count"] == 2

        row = await routes_mod.storage.get_request(rid)
        assert len(row["Закрывающие документы"].splitlines()) == 2

    async def test_someone_elses_request_is_refused(self, api, monkeypatch):
        """Иначе к чужому платежу можно подшить какой угодно документ."""
        client, _ = api
        rid = await self._submit(client, monkeypatch)
        _allow(monkeypatch, "42,77")
        resp = await client.post(
            "/api/my/closing-docs",
            data={"request_id": rid},
            files=[("files", self._doc("akt.pdf"))],
            headers=_auth(77),
        )
        assert resp.status_code == 403

    async def test_unknown_request_is_404(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/my/closing-docs",
            data={"request_id": "INV-20260101-000000-0000"},
            files=[("files", self._doc("akt.pdf"))],
            headers=_auth(),
        )
        assert resp.status_code == 404

    async def test_nothing_attached_is_refused(self, api, monkeypatch):
        client, _ = api
        rid = await self._submit(client, monkeypatch)
        resp = await client.post(
            "/api/my/closing-docs", data={"request_id": rid}, headers=_auth()
        )
        assert resp.status_code == 422

    async def test_financiers_are_told(self, api, monkeypatch):
        """Ждут эти документы в бухгалтерии — молча класть их в таблицу мало."""
        from services import notifier

        client, bot = api
        rid = await self._submit(client, monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555])
        await client.post(
            "/api/my/closing-docs",
            data={"request_id": rid},
            files=[("files", self._doc("akt.pdf"))],
            headers=_auth(),
        )
        sent = " ".join(str(c.kwargs.get("text", "")) for c in bot.send_message.await_args_list)
        assert "Закрывающие документы" in sent
        assert rid in sent
        # Ссылки на каждый документ не перечисляем: безымянные «документ 1,
        # 2, 3» ничего не говорили, а открывают их из строки реестра.
        assert "документ 1" not in sent, sent


class TestOverdueNudge:
    """Автор напоминает финансистам, что его заявка просрочена.

    Планировщик шлёт сводку по всем просроченным сразу, и одна заявка в ней
    теряется; здесь просит человек и про свой конкретный платёж.
    """

    async def _overdue(
        self, client, monkeypatch, when: str = "01.01.2020", who: str = "ООО «Ромашка»"
    ) -> str:
        _allow(monkeypatch)
        resp = await client.post(
            "/api/invoice", data=_form(counterparty=who), headers=_auth()
        )
        assert resp.status_code == 200, resp.text
        rid = resp.json()["request_id"]
        # Просрочку не подать через форму — валидатор не пустит прошедшую
        # дату. Двигаем её в реестре, как двигает её само время.
        await routes_mod.storage.set_request_field(rid, "Плановая дата оплаты", when)
        return rid

    async def test_financiers_get_the_reminder_with_status_buttons(self, api, monkeypatch):
        from services import notifier

        client, bot = api
        rid = await self._overdue(client, monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555, 556])
        bot.send_message.reset_mock()

        resp = await client.post("/api/my/nudge", json={"request_id": rid}, headers=_auth())
        assert resp.status_code == 200, resp.text
        assert resp.json()["ok"] is True

        calls = bot.send_message.await_args_list
        assert [c.kwargs["chat_id"] for c in calls] == [555, 556]
        text = str(calls[0].kwargs["text"])
        assert rid in text and "просрочка" in text
        # Кнопка под напоминанием одна — «Открыть в приложении»: искать
        # карточку месячной давности, чтобы поставить «Оплачено», не надо,
        # а сами статусы теперь ставят в панели.
        assert calls[0].kwargs.get("reply_markup") is not None

    async def test_not_overdue_yet_is_refused(self, api, monkeypatch):
        """Иначе кнопкой можно было бы дёргать по любой заявке."""
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        rid = resp.json()["request_id"]
        resp = await client.post("/api/my/nudge", json={"request_id": rid}, headers=_auth())
        assert resp.status_code == 422

    async def test_paid_request_is_not_overdue(self, api, monkeypatch):
        """Просрочка — это «срок прошёл И всё ещё ждёт»."""
        client, _ = api
        rid = await self._overdue(client, monkeypatch)
        await routes_mod.storage.set_request_field(rid, "Статус оплаты", "Оплачена")
        resp = await client.post("/api/my/nudge", json={"request_id": rid}, headers=_auth())
        assert resp.status_code == 422

    async def test_second_nudge_within_the_window_is_refused(self, api, monkeypatch):
        """Кнопка не должна превращаться в способ забрасывать бухгалтерию."""
        from services import notifier

        client, _ = api
        rid = await self._overdue(client, monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555])
        assert (await client.post(
            "/api/my/nudge", json={"request_id": rid}, headers=_auth()
        )).status_code == 200
        again = await client.post("/api/my/nudge", json={"request_id": rid}, headers=_auth())
        assert again.status_code == 429
        assert "через" in again.json()["detail"]

    async def test_another_request_is_not_blocked_by_the_window(self, api, monkeypatch):
        """Окно на заявку, а не на человека: просрочек может быть несколько."""
        from services import notifier

        client, _ = api
        first = await self._overdue(client, monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555])
        await client.post("/api/my/nudge", json={"request_id": first}, headers=_auth())
        second = await self._overdue(client, monkeypatch, who="ООО «Василёк»")
        resp = await client.post("/api/my/nudge", json={"request_id": second}, headers=_auth())
        assert resp.status_code == 200, resp.text

    async def test_someone_elses_request_is_refused(self, api, monkeypatch):
        client, _ = api
        rid = await self._overdue(client, monkeypatch)
        _allow(monkeypatch, "42,77")
        resp = await client.post("/api/my/nudge", json={"request_id": rid}, headers=_auth(77))
        assert resp.status_code == 403

    async def test_unknown_request_is_404(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.post(
            "/api/my/nudge", json={"request_id": "INV-20260101-000000-0000"},
            headers=_auth(),
        )
        assert resp.status_code == 404

    async def test_undelivered_nudge_frees_the_window(self, api, monkeypatch):
        """Иначе человек шесть часов думал бы, что напоминание ушло."""
        from services import notifier

        client, _ = api
        rid = await self._overdue(client, monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [])
        assert (await client.post(
            "/api/my/nudge", json={"request_id": rid}, headers=_auth()
        )).status_code == 502
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555])
        resp = await client.post("/api/my/nudge", json={"request_id": rid}, headers=_auth())
        assert resp.status_code == 200, resp.text


class TestMiniappLink:
    """Кнопка «Открыть в приложении» под сообщением финансисту.

    В сообщении заявка описана коротко: счёт, реквизиты и историю видно
    только в приложении, а искать там номер руками — лишний шаг.
    """

    def _bot(self, username: str | None = "oplatych_bot"):
        bot = MagicMock()
        # MagicMock отдаёт username как объект, а не строку, — это и есть
        # случай «бот ещё не знает своего имени».
        bot.username = username if username is not None else MagicMock()
        return bot

    def _webapp(self, monkeypatch, short_name: str = "form") -> None:
        monkeypatch.setattr(settings, "webapp_url", "https://pay.example")
        monkeypatch.setattr(settings, "miniapp_short_name", short_name)

    def test_private_chat_gets_a_web_app_button(self, monkeypatch):
        """Главный случай: финансисту пишут в личку.

        web_app открывает приложение сразу и НЕ зависит от короткого имени
        Mini App — прямая ссылка t.me/<бот>/<имя> работает только с именем,
        зарегистрированным в BotFather, и на стенде вела в никуда.
        """
        from services import notifier

        self._webapp(monkeypatch)
        btn = notifier.open_button(self._bot(), "INV-20260101-000000-0001", 555)
        assert btn.web_app is not None
        assert btn.url is None
        assert btn.web_app.url == "https://pay.example?fin=INV-20260101-000000-0001"

    def test_group_gets_a_plain_link(self, monkeypatch):
        """web_app-кнопку Telegram в группах не разрешает — там только ссылка."""
        from services import notifier

        self._webapp(monkeypatch)
        btn = notifier.open_button(self._bot(), "INV-20260101-000000-0001", -100500)
        assert btn.web_app is None
        assert btn.url == (
            "https://t.me/oplatych_bot/form?startapp=fin_INV-20260101-000000-0001"
        )

    def test_group_without_short_name_gets_no_button(self, monkeypatch):
        """Прямую ссылку без имени Mini App не собрать — кнопки не будет."""
        from services import notifier

        self._webapp(monkeypatch, short_name="")
        assert notifier.open_button(self._bot(), "INV-1", -100500) is None
        # А в личке короткое имя не нужно вовсе.
        assert notifier.open_button(self._bot(), "INV-1", 555) is not None

    def test_no_webapp_means_no_button_anywhere(self, monkeypatch):
        from services import notifier

        monkeypatch.setattr(settings, "webapp_url", "")
        assert notifier.open_button(self._bot(), "INV-1", 555) is None

    def test_existing_query_in_webapp_url_is_kept(self, monkeypatch):
        from services import notifier

        monkeypatch.setattr(settings, "webapp_url", "https://pay.example/?v=2")
        btn = notifier.open_button(self._bot(), "INV-1", 555)
        assert btn.web_app.url == "https://pay.example/?v=2&fin=INV-1"

    def test_unknown_bot_username_means_no_link(self, monkeypatch):
        from services import notifier

        self._webapp(monkeypatch)
        assert notifier.miniapp_link(self._bot(None), "INV-1") is None

    def test_card_carries_only_the_open_button(self):
        """Статусы ставят из панели: там они видны вместе со счётом и историей."""
        from telegram import InlineKeyboardButton

        from services import notifier

        extra = InlineKeyboardButton("🔎 Открыть", url="https://t.me/x/y?startapp=z")
        rows = notifier.build_card_keyboard("INV-1", extra).inline_keyboard
        assert [list(r) for r in rows] == [[extra]]

    def test_without_the_app_the_status_buttons_stay(self):
        """Иначе статус ставить будет нечем вообще — карточка обязана давать
        способ действовать."""
        from services import notifier

        rows = notifier.build_card_keyboard("INV-1").inline_keyboard
        assert len(rows) == 1
        assert len(rows[0]) == len(REQUEST_STATUSES)
        assert all(b.callback_data for b in rows[0])

    async def test_nudge_carries_the_button(self, api, monkeypatch):
        from services import notifier

        client, bot = api
        self._webapp(monkeypatch)
        bot.username = "oplatych_bot"
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555])

        _allow(monkeypatch)
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        rid = resp.json()["request_id"]
        await routes_mod.storage.set_request_field(
            rid, "Плановая дата оплаты", "01.01.2020"
        )
        bot.send_message.reset_mock()
        assert (await client.post(
            "/api/my/nudge", json={"request_id": rid}, headers=_auth()
        )).status_code == 200

        markup = bot.send_message.await_args_list[0].kwargs["reply_markup"]
        apps = [b.web_app.url for row in markup.inline_keyboard for b in row if b.web_app]
        assert apps == [f"https://pay.example?fin={rid}"]

    async def test_card_button_matches_each_recipient(self, api, monkeypatch):
        """Одной клавиатурой на всех не обойтись: личка и группа разные."""
        from services import notifier

        client, bot = api
        self._webapp(monkeypatch)
        bot.username = "oplatych_bot"
        # Список получателей проверяется дважды: сначала «настроены ли они
        # вообще» (по .env), потом «кому именно слать» (после резолва).
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "555,-100500")
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555, -100500])
        _allow(monkeypatch)
        bot.send_document.reset_mock()
        bot.send_message.reset_mock()

        assert (await client.post(
            "/api/invoice", data=_form(), headers=_auth()
        )).status_code == 200

        sent = list(bot.send_message.await_args_list) + list(
            bot.send_document.await_args_list
        )
        by_chat = {
            c.kwargs["chat_id"]: c.kwargs.get("reply_markup")
            for c in sent if c.kwargs.get("reply_markup") is not None
        }
        assert by_chat[555].inline_keyboard[0][0].web_app is not None
        assert by_chat[-100500].inline_keyboard[0][0].url.startswith("https://t.me/")


class TestAnimatedHelpFlag:
    """Флаг обкатки живой инструкции: в бою выключен, на стенде включён."""

    async def test_off_by_default(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        resp = await client.get("/api/access", headers=_auth())
        assert resp.json()["animated_help"] is False

    async def test_on_when_enabled(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        monkeypatch.setattr(settings, "animated_help", True)
        resp = await client.get("/api/access", headers=_auth())
        assert resp.json()["animated_help"] is True


class TestAnalyticsRoute:
    """Сводка админа: чужим не показываем, отказ — в аудит."""

    async def test_admin_gets_the_summary(self, api, monkeypatch):
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "42")
        resp = await client.post("/api/invoice", data=_form(), headers=_auth())
        assert resp.status_code == 200

        out = (await client.get("/api/admin/analytics", headers=_auth())).json()
        assert out["flow"]["total_count"] == 1
        assert out["people"]["authors_ever"] == 1
        assert out["docs"]["paid_total"] == 0

    async def test_outsider_is_refused_and_logged(self, api, monkeypatch):
        from services import audit

        client, _ = api
        _allow(monkeypatch, "42,77")
        _admins(monkeypatch, "42")
        resp = await client.get("/api/admin/analytics", headers=_auth(77))
        assert resp.status_code == 403
        events = [e["event"] for e in await audit.recent_events(limit=5)]
        assert audit.ADMIN_DENIED in events

    async def test_period_is_clamped(self, api, monkeypatch):
        """days из адресной строки — чужой ввод: без границ он уедет в вечность."""
        client, _ = api
        _allow(monkeypatch)
        _admins(monkeypatch, "42")
        for asked, got in ((0, 1), (5000, 365), (30, 30)):
            out = (await client.get(
                f"/api/admin/analytics?days={asked}", headers=_auth()
            )).json()
            assert out["days"] == got


class TestAssetVersioning:
    """Адрес файла меняется вместе с его содержимым.

    Без этого WebView Telegram неделями показывал старый JS: заголовок
    no-cache обязывает переспросить сервер, но страницу WebView держит
    живой, и человек после деплоя видел прежний экран.
    """

    def test_local_assets_get_a_version_and_the_cdn_one_does_not(self):
        from api.server import asset_version, index_html

        page = index_html(asset_version())
        assert 'src="app.js?v=' in page
        assert 'href="app.css?v=' in page
        # Ссылка на telegram.org — чужая, версию туда дописывать нельзя.
        assert 'src="https://telegram.org/js/telegram-web-app.js"' in page
        assert "telegram-web-app.js?v=" not in page

    def test_version_follows_the_content(self, tmp_path, monkeypatch):
        import api.server as server_mod

        room = tmp_path / "webapp"
        room.mkdir()
        (room / "index.html").write_text('<script src="a.js"></script>', encoding="utf-8")
        (room / "a.js").write_text("var a = 1;", encoding="utf-8")
        monkeypatch.setattr(server_mod, "WEBAPP_DIR", room)

        first = server_mod.asset_version()
        assert server_mod.asset_version() == first, "версия скачет без правок"
        (room / "a.js").write_text("var a = 2;", encoding="utf-8")
        assert server_mod.asset_version() != first, "правку файла версия не заметила"

    async def test_page_is_served_with_versions(self, api):
        client, _ = api
        page = (await client.get("/")).text
        assert "?v=" in page and 'src="app.js?v=' in page


class TestSilenceSwitch:
    """Тумблер «не присылать мне ничего» — один на все потоки сообщений."""

    def _setup(self, monkeypatch, ids: str = "42"):
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", ids)

    async def test_silence_stops_cards_of_new_requests(self, api, monkeypatch):
        from services import notifier

        client, bot = api
        self._setup(monkeypatch, "555")
        _allow(monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [555])
        rs.set_personal_reminders(555, silent=True)
        bot.send_message.reset_mock()
        bot.send_document.reset_mock()

        assert (await client.post(
            "/api/invoice", data=_form(), headers=_auth()
        )).status_code == 200
        to_financier = [
            c for c in list(bot.send_message.await_args_list)
            + list(bot.send_document.await_args_list)
            if c.kwargs.get("chat_id") == 555
        ]
        assert to_financier == [], "карточка ушла тому, кто просил тишины"
        # Автору подтверждение приходит по-прежнему: тишину включил не он.
        assert any(
            c.kwargs.get("chat_id") == 42
            for c in list(bot.send_message.await_args_list)
            + list(bot.send_document.await_args_list)
        ), "автор остался без подтверждения"

    async def test_silence_beats_urgent(self, api, monkeypatch):
        """Срочные приходят всегда — но не тому, кто выключил всё."""
        from services import notifier
        from tests.conftest import make_request

        self._setup(monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [42])
        rs.set_personal_reminders(42, silent=True)
        urgent = make_request(urgency=__import__(
            "bot.models", fromlist=["Urgency"]).Urgency.URGENT)
        assert notifier.recipients_for(urgent) == []

    async def test_silence_stops_reminders_even_on_manual_run(self, tmp_paths, monkeypatch):
        from unittest.mock import AsyncMock, MagicMock

        from services import reminders, storage
        from tests.conftest import make_request

        self._setup(monkeypatch, "111")
        await storage.append_invoice(make_request(
            planned_date=date(2026, 8, 5), request_id="INV-20260804-100030-0030"
        ))
        rs.set_personal_reminders(111, silent=True)
        bot = MagicMock()
        bot.send_message = AsyncMock()
        rows = await storage.recent_requests(limit=50)
        assert await reminders.send_to(bot, 111, rows, date(2026, 8, 4), force=True) == (0, 0)
        bot.send_message.assert_not_awaited()

    async def test_turning_it_back_on_restores_delivery(self, api, monkeypatch):
        from services import notifier
        from tests.conftest import make_request

        self._setup(monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [42])
        rs.set_personal_reminders(42, silent=True)
        assert notifier.recipients_for(make_request()) == []
        rs.set_personal_reminders(42, silent=False)
        assert notifier.recipients_for(make_request()) == [42]

    async def test_last_recipient_going_silent_is_warned(self, api, monkeypatch):
        """Замолчали все — о новых заявках не узнает никто. Сказать об этом."""
        from services import notifier

        client, _ = api
        self._setup(monkeypatch)
        _allow(monkeypatch)
        monkeypatch.setattr(notifier, "resolved_finance_ids", lambda: [42])
        resp = await client.post("/api/reminders/me", json={"silent": True}, headers=_auth())
        assert resp.status_code == 200
        assert "не узнает никто" in resp.json()["message"]
