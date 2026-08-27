"""e2e чат-формы: прогон шагов диалога до финализации, без Telegram.

Мок — только транспорт (reply_text/бот); валидация, дедуп, реестр и PDF —
настоящие.
"""
from __future__ import annotations

from unittest.mock import AsyncMock, MagicMock

from openpyxl import load_workbook
from telegram.ext import ConversationHandler

import bot.handlers as h
from bot.scheduling import auto_planned_date
from config import settings


def _bot() -> MagicMock:
    bot = MagicMock()
    bot.send_message = AsyncMock()
    bot.send_document = AsyncMock()
    member = MagicMock()
    member.status = "member"
    bot.get_chat_member = AsyncMock(return_value=member)
    return bot


def _ctx(bot: MagicMock) -> MagicMock:
    ctx = MagicMock()
    ctx.user_data = {}
    ctx.bot = bot
    return ctx


def _user(update: MagicMock, user_id: int) -> None:
    update.effective_user.id = user_id
    update.effective_user.username = "tester"
    update.effective_user.full_name = "Тест Тестов"
    update.effective_chat.type = "private"


def _msg(text: str, user_id: int = 42) -> MagicMock:
    u = MagicMock()
    u.message.text = text
    u.message.reply_text = AsyncMock()
    u.effective_message = u.message
    _user(u, user_id)
    return u


def _cb(data: str, user_id: int = 42) -> MagicMock:
    u = MagicMock()
    q = u.callback_query
    q.data = data
    q.answer = AsyncMock()
    q.edit_message_text = AsyncMock()
    q.message.reply_text = AsyncMock()
    u.effective_message = q.message
    _user(u, user_id)
    return u


async def _drive_common(ctx, urgency_cb: str) -> None:
    assert await h.step_amount(_msg("125 000,50"), ctx) == h.CURRENCY
    assert await h.step_currency(_cb("CUR:RUB"), ctx) == h.COUNTERPARTY
    assert await h.step_counterparty(_msg("ООО «Ромашка»"), ctx) == h.ARTICLE
    result = await h.step_article(_cb("ART:0"), ctx)
    assert result == h.URGENCY
    assert await h.step_urgency_choice(_cb(urgency_cb), ctx) in (
        h.WORK_DEADLINE, h.PLANNED_DATE
    )
    # Срок исполнения работ — шаг между датой оплаты и комментарием.
    if ctx.user_data.get(h.K_PLANNED) is not None:
        assert await h.step_work_deadline(_msg("текущий месяц"), ctx) == h.COMMENT


async def test_happy_path_requisites(tmp_paths):
    bot = _bot()
    ctx = _ctx(bot)
    await _drive_common(ctx, "URG:NORMAL")
    assert await h.step_comment(_msg("аренда за июль"), ctx) == h.INVOICE_CHOICE
    assert await h.step_invoice_choice(_cb("INV_NO"), ctx) == h.REQUISITES
    # По ТЗ: запись — только после подтверждения.
    # Шаг дополнительных документов необязателен, но он в диалоге есть:
    # договор бывает и у заявки по реквизитам.
    assert await h.step_requisites(_msg("ИНН 7707083893"), ctx) == h.EXTRA_DOCS
    assert await h.extra_docs_done(_cb("EXTRA_DONE"), ctx) == h.CONFIRM_SUBMIT
    assert await h.submit_confirm(_cb("SUB_YES"), ctx) == ConversationHandler.END

    assert ctx.user_data == {}  # диалог очищен
    ws = load_workbook(settings.registry_path).active
    assert ws.max_row == 2
    assert ws.cell(2, 4).value == "ООО «Ромашка»"
    # Дата — серверная: следующий рабочий день для «Обычной».
    assert ws.cell(2, 2).value == auto_planned_date(False).strftime("%d.%m.%Y")
    bot.send_document.assert_awaited()  # подтверждение автору (PDF)


async def test_custom_date_path(tmp_paths):
    ctx = _ctx(_bot())
    await _drive_common(ctx, "URG:CUSTOM")
    assert await h.step_planned_date(_msg("31.12.2026"), ctx) == h.WORK_DEADLINE
    assert await h.step_work_deadline(_msg("поставка в декабре"), ctx) == h.COMMENT
    assert await h.step_comment(_msg("оплата по договору"), ctx) == h.INVOICE_CHOICE
    assert await h.step_invoice_choice(_cb("INV_NO"), ctx) == h.REQUISITES
    assert await h.step_requisites(_msg("реквизиты"), ctx) == h.EXTRA_DOCS
    assert await h.extra_docs_done(_cb("EXTRA_DONE"), ctx) == h.CONFIRM_SUBMIT
    assert await h.submit_confirm(_cb("SUB_YES"), ctx) == ConversationHandler.END

    ws = load_workbook(settings.registry_path).active
    assert ws.cell(2, 2).value == "31.12.2026"
    assert ws.cell(2, 11).value == "Обычная"  # настраиваемая дата ≠ срочность


async def test_comment_can_be_skipped(tmp_paths):
    ctx = _ctx(_bot())
    await _drive_common(ctx, "URG:NORMAL")
    assert await h.comment_skip(_cb("CMT_SKIP"), ctx) == h.INVOICE_CHOICE
    assert ctx.user_data[h.K_COMMENT] == ""
    assert await h.step_invoice_choice(_cb("INV_NO"), ctx) == h.REQUISITES
    assert await h.step_requisites(_msg("реквизиты без комментария"), ctx) \
        == h.EXTRA_DOCS
    assert await h.extra_docs_done(_cb("EXTRA_DONE"), ctx) == h.CONFIRM_SUBMIT
    assert await h.submit_confirm(_cb("SUB_YES"), ctx) == ConversationHandler.END

    ws = load_workbook(settings.registry_path).active
    assert ws.cell(2, 8).value in (None, "")  # колонка «Комментарий» пуста


async def test_duplicate_requires_confirmation(tmp_paths):
    bot = _bot()

    async def submit(ctx):
        await _drive_common(ctx, "URG:NORMAL")
        await h.step_comment(_msg("аренда за июль"), ctx)
        await h.step_invoice_choice(_cb("INV_NO"), ctx)
        assert await h.step_requisites(_msg("ИНН 7707083893"), ctx) == h.EXTRA_DOCS
        assert await h.extra_docs_done(_cb("EXTRA_DONE"), ctx) == h.CONFIRM_SUBMIT
        return await h.submit_confirm(_cb("SUB_YES"), ctx)

    assert await submit(_ctx(bot)) == ConversationHandler.END

    ctx2 = _ctx(bot)
    assert await submit(ctx2) == h.DUP_CONFIRM        # дубль пойман после подтверждения
    assert h.K_PENDING_REQ in ctx2.user_data
    assert await h.dup_confirm(_cb("DUP_YES"), ctx2) == ConversationHandler.END

    ws = load_workbook(settings.registry_path).active
    assert ws.max_row == 3  # обе заявки в реестре

    ctx3 = _ctx(bot)
    assert await submit(ctx3) == h.DUP_CONFIRM
    assert await h.dup_confirm(_cb("DUP_NO"), ctx3) == ConversationHandler.END
    ws = load_workbook(settings.registry_path).active
    assert ws.max_row == 3  # отказ — ничего не записано


async def test_confirmation_cancel_writes_nothing(tmp_paths):
    ctx = _ctx(_bot())
    await _drive_common(ctx, "URG:NORMAL")
    await h.step_comment(_msg("проверка отмены"), ctx)
    await h.step_invoice_choice(_cb("INV_NO"), ctx)
    assert await h.step_requisites(_msg("реквизиты"), ctx) == h.EXTRA_DOCS
    assert await h.extra_docs_done(_cb("EXTRA_DONE"), ctx) == h.CONFIRM_SUBMIT
    assert await h.submit_confirm(_cb("SUB_NO"), ctx) == ConversationHandler.END
    assert ctx.user_data == {}
    assert not settings.registry_path.exists()  # ничего не записано


async def test_neither_invoice_nor_requisites(tmp_paths):
    """Третий вариант выбора: ни счёта, ни реквизитов.

    Раньше диалог требовал одно из двух, и заявку «оплатить по договору,
    документы будут позже» подать было нельзя — вписывали что попало.
    """
    bot = _bot()
    ctx = _ctx(bot)
    await _drive_common(ctx, "URG:NORMAL")
    assert await h.step_comment(_msg("по договору"), ctx) == h.INVOICE_CHOICE
    assert await h.step_invoice_choice(_cb("INV_NONE"), ctx) == h.EXTRA_DOCS
    assert await h.extra_docs_done(_cb("EXTRA_DONE"), ctx) == h.CONFIRM_SUBMIT
    assert await h.submit_confirm(_cb("SUB_YES"), ctx) == ConversationHandler.END

    ws = load_workbook(settings.registry_path).active
    assert ws.max_row == 2
    from bot.models import SHEET_HEADERS
    assert (ws.cell(2, SHEET_HEADERS.index("Ссылка на счет") + 1).value or "") == ""
    assert (ws.cell(2, SHEET_HEADERS.index("Реквизиты") + 1).value or "") == ""


async def test_financier_is_told_there_is_nothing_to_pay_against(tmp_paths):
    """Пустая заявка законна, но карточка обязана это назвать."""
    from decimal import Decimal

    from bot.models import InvoiceRequest, Urgency
    from services.notifier import _format_card

    card = _format_card(
        InvoiceRequest(
            telegram_id=1, sender_username="@t", sender_name="Т",
            amount=Decimal("100"), currency="RUB", counterparty="ООО",
            comment="", urgency=Urgency.NORMAL, has_invoice=False, requisites="",
        ),
        row_number=1,
    )
    assert "Ни счёта, ни реквизитов" in card
    assert "Счёт — этим файлом" not in card
