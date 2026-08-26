"""xlsx-реестр: запись, нумерация, статусы, подсказки, защита от формул."""
from __future__ import annotations

from openpyxl import load_workbook

from bot.models import SHEET_HEADERS
from config import settings
from services.registry_xlsx import (
    append_sync,
    recent_counterparties_sync,
    set_status_sync,
)
from tests.conftest import make_request


def _reg():
    return settings.registry_path


def test_append_numbering_and_headers(tmp_paths):
    assert append_sync(make_request(telegram_id=1), _reg()) == 1
    assert append_sync(make_request(telegram_id=2, comment="много\nстрок"), _reg()) == 2
    assert append_sync(make_request(telegram_id=3), _reg()) == 3

    ws = load_workbook(_reg()).active
    header = [c.value for c in ws[1]]
    assert header[: len(SHEET_HEADERS)] == SHEET_HEADERS
    assert ws.max_row == 4  # шапка + три заявки


def test_column_order_is_frozen(tmp_paths):
    """Порядок колонок зафиксирован, новые приписываются только в конец.

    Раньше тест держал одно: «последняя — срок работ». Он и упал, когда в
    конец легли «Дополнительные документы» — то есть сторожил не то. Вставка
    в СЕРЕДИНУ сдвинула бы все заполненные строки на колонку влево, поэтому
    проверяем весь список целиком: любое перемещение видно сразу, а
    дописывание в конец требует осознанно поправить эту строку.
    """
    assert SHEET_HEADERS == [
        "Дата внесения в реестр", "Плановая дата оплаты", "Сотрудник по заявке",
        "Контрагент", "Сумма", "Статья", "Статус оплаты", "Комментарий",
        "Ссылка на счет",                       # первые девять — ровно по ТЗ
        "Валюта", "Срочность", "Реквизиты", "ID заявки", "Telegram ID",
        "Срок исполнения работ по договору",
        "Дополнительные документы",
    ]
    assert append_sync(make_request(work_deadline="услуга на 6 месяцев"), _reg()) == 1
    ws = load_workbook(settings.registry_path).active
    deadline_col = SHEET_HEADERS.index("Срок исполнения работ по договору") + 1
    assert ws.cell(1, deadline_col).value == "Срок исполнения работ по договору"
    assert ws.cell(2, deadline_col).value == "услуга на 6 месяцев"


def test_work_deadline_may_be_empty(tmp_paths):
    """Заявки, поданные до появления поля, читаются и пишутся без него.

    В форме поле обязательное, но в модели остаётся пустая строка — иначе
    прежние строки реестра перестали бы открываться.
    """
    append_sync(make_request(), _reg())
    ws = load_workbook(settings.registry_path).active
    assert (ws.cell(2, len(SHEET_HEADERS)).value or "") == ""


def test_tz_columns_come_first(tmp_paths):
    """Первые девять колонок — ровно по ТЗ (сверено с шаблоном Реестр.xlsx)."""
    assert SHEET_HEADERS[:9] == [
        "Дата внесения в реестр",
        "Плановая дата оплаты",
        "Сотрудник по заявке",
        "Контрагент",
        "Сумма",
        "Статья",
        "Статус оплаты",
        "Комментарий",
        "Ссылка на счет",
    ]
    row = make_request().as_sheet_row()
    assert len(row) == len(SHEET_HEADERS)
    assert row[1] == "15.08.2026"      # плановая дата
    assert row[5] == "Аренда"          # статья
    assert row[6] == "Новая"           # статус оплаты
    assert row[8] == ""                # нет счёта → ссылка пустая (по ТЗ)


def test_set_status_returns_row(tmp_paths):
    r = make_request(telegram_id=777)
    append_sync(r, _reg())
    row = set_status_sync(r.request_id, "Оплачена", _reg())
    assert row is not None
    assert row["Статус оплаты"] == "Оплачена"
    assert row["Telegram ID"] == "777"
    assert set_status_sync("INV-00000000-000000-0000", "Оплачена", _reg()) is None
    # Реестр остаётся валидным, нумерация продолжается.
    assert append_sync(make_request(telegram_id=778), _reg()) == 2


def test_recent_counterparties_order(tmp_paths):
    append_sync(make_request(telegram_id=1, counterparty="ООО «Ромашка»"), _reg())
    append_sync(make_request(telegram_id=2, counterparty="ИП Иванов"), _reg())
    append_sync(make_request(telegram_id=3, counterparty="ООО «Ромашка»"), _reg())
    append_sync(make_request(telegram_id=4, counterparty="ЗАО Тест"), _reg())
    top = recent_counterparties_sync(6, _reg())
    assert top[0] == "ООО «Ромашка»"
    assert set(top) == {"ООО «Ромашка»", "ИП Иванов", "ЗАО Тест"}


def test_formula_injection_escaped(tmp_paths):
    r = make_request(counterparty="=IMPORTRANGE(\"evil\";\"A1\")")
    append_sync(r, _reg())
    ws = load_workbook(_reg()).active
    cell = ws.cell(2, SHEET_HEADERS.index("Контрагент") + 1)
    assert str(cell.value).startswith("'=")  # экранировано, не формула
    assert cell.data_type != "f"


def test_csv_registry_name_maps_to_xlsx(tmp_paths, monkeypatch):
    """Бэккомпат: старое REGISTRY_FILE=registry.csv превращается в .xlsx."""
    monkeypatch.setattr(settings, "registry_file", "registry.csv")
    assert settings.registry_path.suffix == ".xlsx"


def test_two_writers_survive_a_new_column(tmp_paths):
    """Гонка миграции: обе одновременные подачи должны пройти.

    _add_missing_columns смотрит PRAGMA и делает ALTER. Две подачи в один
    момент обе видят, что колонки нет, и обе идут её создавать — вторая
    получала «duplicate column name» и роняла заявку с 500. Ловилось
    нестабильно: только в первый запуск после добавления колонки.
    """
    import sqlite3

    from config import settings
    from services import registry_sqlite as reg

    # Таблица «из прошлой версии»: без последней колонки.
    conn = sqlite3.connect(settings.security_db_path)
    older = [f for f in reg._FIELDS if f != "extra_files"]
    conn.execute(
        "CREATE TABLE IF NOT EXISTS requests "
        "(id INTEGER PRIMARY KEY AUTOINCREMENT, ts REAL, "
        + ", ".join(f"{f} TEXT" for f in older) + ")"
    )
    conn.commit()

    first = sqlite3.connect(settings.security_db_path)
    second = sqlite3.connect(settings.security_db_path)
    try:
        reg._add_missing_columns(first)
        first.commit()
        # Второе соединение уже прочитало старую схему — повторный ALTER.
        reg._add_missing_columns(second)
        second.commit()
    finally:
        first.close()
        second.close()
        conn.close()

    assert reg.append_sync(make_request()) == 1
