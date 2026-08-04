"""Безопасность файлов и xlsx-зеркало."""
from __future__ import annotations

from decimal import Decimal

from services.local_storage import _save_sync, build_invoice_filename
from tests.conftest import NOW, make_request


class TestFilenameSanitization:
    def test_normal(self):
        name = build_invoice_filename("счёт.pdf", "ООО «Ромашка»", Decimal("125000.50"), NOW)
        # Decimal + ".0f" округляет по банковскому правилу: 125000.50 → 125000.
        assert name == "20260803_ООО Ромашка_125000.pdf"

    def test_path_traversal_in_extension_blocked(self):
        name = build_invoice_filename(
            "invoice.p/../../etc/passwd", "ООО Тест", Decimal("100"), NOW
        )
        assert "/" not in name and ".." not in name

    def test_path_traversal_in_counterparty_blocked(self):
        name = build_invoice_filename("a.pdf", "../../etc", Decimal("100"), NOW)
        assert "/" not in name and ".." not in name

    def test_empty_parts_get_fallbacks(self):
        name = build_invoice_filename("noext", "***", Decimal("1"), NOW)
        assert name == "20260803_invoice_1.bin"


def test_save_collision_adds_suffix(tmp_paths):
    p1 = _save_sync(content=b"one", filename="a.pdf")
    p2 = _save_sync(content=b"two", filename="a.pdf")
    assert p1 != p2
    assert p1.read_bytes() == b"one"
    assert p2.read_bytes() == b"two"
    assert p2.name == "a_1.pdf"


def test_xlsx_append_and_status(tmp_paths):
    from openpyxl import load_workbook

    from services import registry_xlsx

    path = tmp_paths / "Реестр.xlsx"
    r = make_request(telegram_id=555)
    registry_xlsx.append_sync(r, path)
    registry_xlsx.set_status_sync(r.request_id, "Оплачена", path)

    ws = load_workbook(path).active
    assert ws.cell(1, 1).value == "Дата внесения в реестр"
    assert ws.cell(2, 4).value == "ООО «Ромашка»"   # Контрагент
    assert ws.cell(2, 7).value == "Оплачена"        # Статус оплаты


def test_xlsx_respects_tz_template(tmp_paths):
    """Лист с 9 колонками из шаблона ТЗ дополняется служебными заголовками."""
    from openpyxl import Workbook, load_workbook

    from bot.models import SHEET_HEADERS
    from services import registry_xlsx

    path = tmp_paths / "Реестр.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(SHEET_HEADERS[:9])  # как в выданном шаблоне
    wb.save(path)

    registry_xlsx.append_sync(make_request(), path)

    ws = load_workbook(path).active
    assert ws.cell(1, 10).value == "Валюта"       # служебный заголовок дописан
    assert ws.cell(2, 2).value == "15.08.2026"    # плановая дата в строке
