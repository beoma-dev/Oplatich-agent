"""Горизонт 3: SQLite-реестр, синхронизация карточек, причина статуса, миграция."""
from __future__ import annotations

from unittest.mock import AsyncMock, MagicMock

import pytest
from telegram.error import Forbidden, NetworkError
from telegram.ext import ApplicationHandlerStop

import bot.finance_actions as fa
from bot.models import Urgency
from config import settings
from services import (
    cards,
    intake,
    notifier,
    registry_sqlite,
    registry_xlsx,
    storage,
    tg_retry,
)
from tests.conftest import make_request


class TestSqliteRegistry:
    def test_append_numbering_and_unique(self, tmp_paths):
        r1, r2 = make_request(telegram_id=1), make_request(telegram_id=2)
        assert registry_sqlite.append_sync(r1) == 1
        assert registry_sqlite.append_sync(r2) == 2
        assert not registry_sqlite.import_row_sync(r1.as_sheet_row())  # дубль ID

    def test_work_deadline_roundtrip(self, tmp_paths):
        r = make_request(telegram_id=555, work_deadline="услуга на 6 месяцев")
        registry_sqlite.append_sync(r)
        row = registry_sqlite.get_request_sync(r.request_id)
        assert row is not None
        assert row["Срок исполнения работ по договору"] == "услуга на 6 месяцев"

    def test_missing_column_is_added_to_existing_table(self, tmp_paths):
        """Боевая БД создана до появления колонки — INSERT не должен падать.

        CREATE TABLE IF NOT EXISTS готовую таблицу не трогает, поэтому у уже
        работающего бота колонка сама не появится: её досоздаёт миграция.
        Воспроизводим ровно это — таблица «из прошлого», без новой колонки.
        """
        import sqlite3

        path = settings.security_db_path
        path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(path) as conn:
            conn.execute(
                """
                CREATE TABLE requests (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ts REAL NOT NULL,
                    created_at TEXT, planned_date TEXT, sender TEXT,
                    counterparty TEXT, amount TEXT, article TEXT, status TEXT,
                    comment TEXT, file_link TEXT, currency TEXT, urgency TEXT,
                    requisites TEXT, request_id TEXT UNIQUE NOT NULL, telegram_id TEXT
                )
                """
            )
        r = make_request(telegram_id=556, work_deadline="поставка в декабре")
        assert registry_sqlite.append_sync(r) == 1
        row = registry_sqlite.get_request_sync(r.request_id)
        assert row["Срок исполнения работ по договору"] == "поставка в декабре"

    def test_set_status_roundtrip(self, tmp_paths):
        r = make_request(telegram_id=777)
        registry_sqlite.append_sync(r)
        row = registry_sqlite.set_status_sync(r.request_id, "Оплачена")
        assert row is not None
        assert row["Статус оплаты"] == "Оплачена"
        assert row["Telegram ID"] == "777"
        assert registry_sqlite.set_status_sync("INV-00000000-000000-0000", "x") is None

    def test_counterparties(self, tmp_paths):
        for uid, cp in [(1, "А"), (2, "Б"), (3, "А")]:
            registry_sqlite.append_sync(make_request(telegram_id=uid, counterparty=cp))
        assert registry_sqlite.recent_counterparties_sync(5) == ["А", "Б"]


class TestStorageFacadeLocal:
    async def test_append_writes_primary_and_mirror(self, tmp_paths):
        r = make_request()
        assert await storage.append_invoice(r) == 1
        assert registry_sqlite.has_request_sync(r.request_id)      # первичное
        assert settings.registry_path.exists()                      # xlsx-зеркало

    async def test_status_updates_both(self, tmp_paths):
        from openpyxl import load_workbook

        r = make_request()
        await storage.append_invoice(r)
        row = await storage.set_request_status(r.request_id, "Отклонена")
        assert row is not None and row["Статус оплаты"] == "Отклонена"
        ws = load_workbook(settings.registry_path).active
        assert ws.cell(2, 7).value == "Отклонена"                   # зеркало догнало


class TestCardsStore:
    async def test_roundtrip_and_replace(self, tmp_paths):
        await cards.save("INV-1", 10, 100, True, "base")
        await cards.save("INV-1", 20, 200, False, "base2")
        await cards.save("INV-1", 10, 100, True, "base-updated")    # replace
        got = await cards.for_request("INV-1")
        assert len(got) == 2
        assert {c["chat_id"] for c in got} == {10, 20}


class TestReasonFlow:
    def _bot(self):
        bot = MagicMock()
        message = MagicMock()
        message.chat.id = 555
        message.message_id = 77
        bot.send_document = AsyncMock(return_value=message)
        bot.send_message = AsyncMock(return_value=message)
        bot.edit_message_caption = AsyncMock()
        bot.edit_message_text = AsyncMock()
        return bot

    def _cb(self, data, user_id=555):
        u = MagicMock()
        q = u.callback_query
        q.data = data
        q.answer = AsyncMock()
        q.edit_message_text = AsyncMock()
        q.message.chat.id = user_id
        q.message.message_id = 77
        q.message.caption = "карточка"
        q.message.caption_html = "карточка"
        u.effective_user.id = user_id
        u.effective_user.username = "fin"
        u.effective_user.full_name = "Финансист"
        return u

    def _ctx(self, bot):
        ctx = MagicMock()
        ctx.bot = bot
        return ctx

    async def test_reject_asks_reason_then_applies_everywhere(self, tmp_paths, monkeypatch):
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "555")
        monkeypatch.setattr(fa, "_pending_reasons", {})
        bot = self._bot()

        r = make_request(telegram_id=901)
        await storage.append_invoice(r)
        assert await notifier.notify_finance(bot, r, 1, pdf=b"%PDF") == 1  # карточка сохранена

        # Клик «Отклонено» — статус НЕ меняется, бот просит причину.
        await fa.status_callback(self._cb(f"ST:{r.request_id}:REJECTED"), self._ctx(bot))
        assert 555 in fa._pending_reasons
        prompt = bot.send_message.call_args.kwargs["text"]
        assert "причину" in prompt.lower()
        row = registry_sqlite.set_status_sync(r.request_id, "Новая")  # ещё «Новая»
        assert row is not None

        # Причина сообщением → статус, карточка у финансиста, автору — причина.
        upd = MagicMock()
        upd.effective_user.id = 555
        upd.effective_user.username = "fin"
        upd.effective_user.full_name = "Финансист"
        upd.effective_message.text = "нет счёта-фактуры"
        upd.effective_message.reply_text = AsyncMock()
        with pytest.raises(ApplicationHandlerStop):
            await fa.reason_message(upd, self._ctx(bot))

        assert registry_sqlite.set_status_sync(r.request_id, "Отклонена") is not None
        caption = bot.edit_message_caption.call_args.kwargs["caption"]
        assert "Отклонено" in caption and "нет счёта-фактуры" in caption
        author_msg = bot.send_message.call_args.kwargs["text"]
        assert bot.send_message.call_args.kwargs["chat_id"] == 901  # уведомлён автор
        assert "нет счёта-фактуры" in author_msg
        assert "Отклонено" in author_msg

    async def test_paid_applies_immediately(self, tmp_paths, monkeypatch):
        settings.__dict__.pop("finance_recipients", None)
        monkeypatch.setattr(settings, "finance_chat_ids_raw", "555")
        monkeypatch.setattr(fa, "_pending_reasons", {})
        bot = self._bot()
        r = make_request(telegram_id=902)
        await storage.append_invoice(r)
        await notifier.notify_finance(bot, r, 1, pdf=b"%PDF")

        await fa.status_callback(self._cb(f"ST:{r.request_id}:PAID"), self._ctx(bot))
        assert fa._pending_reasons == {}
        row = await storage.set_request_status(r.request_id, "Оплачена")
        assert row is not None
        assert "Оплачено" in bot.edit_message_caption.call_args.kwargs["caption"]

    async def test_stray_text_passes_through(self, tmp_paths, monkeypatch):
        monkeypatch.setattr(fa, "_pending_reasons", {})
        upd = MagicMock()
        upd.effective_user.id = 42
        upd.effective_message.text = "просто сообщение"
        # Нет pending — хендлер молчит и НЕ останавливает обработку.
        await fa.reason_message(upd, self._ctx(self._bot()))


class TestMigration:
    def test_xlsx_rows_move_to_sqlite(self, tmp_paths):
        from scripts.migrate_xlsx_to_sqlite import migrate

        r1, r2 = make_request(telegram_id=1), make_request(telegram_id=2)
        registry_xlsx.append_sync(r1, settings.registry_path)
        registry_xlsx.append_sync(r2, settings.registry_path)

        imported, skipped = migrate()
        assert (imported, skipped) == (2, 0)
        assert registry_sqlite.has_request_sync(r1.request_id)
        imported2, skipped2 = migrate()   # повторный запуск безопасен
        assert (imported2, skipped2) == (0, 2)


class TestStaleCards:
    """Необновлённая карточка врёт, и об этом должно быть слышно.

    Статус в реестре уже сменён; карточка со старым статусом держит живые
    кнопки, и второй финансист нажимает по ней ещё раз. Раньше сетевой сбой
    обновления гасился на уровне DEBUG — то есть не был виден вообще.
    """

    async def test_network_failure_raises_an_alert(self, tmp_paths, monkeypatch):
        sent = {}

        async def fake_alert(bot, title, details="", **kw):
            sent["title"], sent["kind"] = title, kw.get("kind")
            return 1

        monkeypatch.setattr(cards.alerts, "alert_admins", fake_alert)
        monkeypatch.setattr(cards, "for_request", AsyncMock(return_value=[
            {"chat_id": 1, "message_id": 2, "is_caption": False, "base_html": "x"},
        ]))
        bot = MagicMock()
        bot.edit_message_text = AsyncMock(side_effect=NetworkError("канал молчит"))

        assert await cards.update_all(bot, "INV-1", " · статус") == 0
        assert "старым статусом" in sent["title"]
        assert sent["kind"] == "delivery", "своей категории для этого заводить незачем"

    async def test_deleted_card_stays_quiet(self, tmp_paths, monkeypatch):
        """Сообщение удалили или оно не изменилось — это не сбой канала."""
        called = []
        monkeypatch.setattr(cards.alerts, "alert_admins",
                            AsyncMock(side_effect=lambda *a, **k: called.append(1)))
        monkeypatch.setattr(cards, "for_request", AsyncMock(return_value=[
            {"chat_id": 1, "message_id": 2, "is_caption": False, "base_html": "x"},
        ]))
        bot = MagicMock()
        bot.edit_message_text = AsyncMock(side_effect=RuntimeError("message to edit not found"))

        assert await cards.update_all(bot, "INV-1", " · статус") == 0
        assert called == [], "об удалённой карточке будить незачем"


class TestCardRetry:
    """Повтор отправки карточки на сетевых сбоях.

    Канал до Telegram идёт через WARP и пропадает на минуты. Заявка уже
    в реестре, а карточку потом никто не перешлёт — однократный сбой сети
    не должен её стоить.

    Сам механизм с 25.08.2026 живёт в services/tg_retry и общий для всех
    исходящих; зовём его через notifier — так проверяется и проводка.
    """

    async def test_second_attempt_succeeds(self, monkeypatch):
        monkeypatch.setattr(tg_retry.asyncio, "sleep", AsyncMock())
        calls = []

        async def flaky():
            calls.append(1)
            if len(calls) == 1:
                raise NetworkError("proxy unreachable")
            return "ok"

        assert await notifier._send_with_retry(flaky, 42) == "ok"
        assert len(calls) == 2

    async def test_gives_up_after_the_budget(self, monkeypatch):
        monkeypatch.setattr(tg_retry.asyncio, "sleep", AsyncMock())
        calls = []

        async def always_down():
            calls.append(1)
            raise NetworkError("proxy unreachable")

        with pytest.raises(NetworkError):
            await notifier._send_with_retry(always_down, 42)
        assert len(calls) == len(tg_retry.DEFAULT_PAUSES) + 1

    async def test_meaningful_refusals_are_not_retried(self, monkeypatch):
        """Бот заблокирован или чат не найден — повторять бессмысленно."""
        monkeypatch.setattr(tg_retry.asyncio, "sleep", AsyncMock())
        calls = []

        async def forbidden():
            calls.append(1)
            raise Forbidden("bot was blocked by the user")

        with pytest.raises(Forbidden):
            await notifier._send_with_retry(forbidden, 42)
        assert len(calls) == 1


class TestNotifications:
    """Кому что уходит после подачи заявки."""

    @staticmethod
    def _bot():
        bot = MagicMock()
        bot.send_message = AsyncMock()
        bot.send_document = AsyncMock()
        # Итог в группу публикуется только участнику чата: return_chat_id
        # приходит из ссылки и может быть подделан.
        member = MagicMock()
        member.status = "member"
        bot.get_chat_member = AsyncMock(return_value=member)
        return bot

    async def test_group_summary_carries_the_comment(self, tmp_paths):
        bot = self._bot()
        request = make_request(comment="аренда офиса за август", work_deadline="текущий месяц")
        await intake._post_group_summary(bot, request, -100123)
        text = bot.send_message.await_args.kwargs["text"]
        assert "аренда офиса за август" in text
        assert "текущий месяц" in text

    async def test_group_summary_shows_a_dash_without_comment(self, tmp_paths):
        bot = self._bot()
        await intake._post_group_summary(bot, make_request(comment=""), -100123)
        assert "Комментарий: —" in bot.send_message.await_args.kwargs["text"]

    async def test_author_gets_nothing_when_the_group_already_knows(self):
        """Итог уже в группе — вторым сообщением тот же текст автору не пишем."""
        bot = self._bot()
        await intake._send_user_confirmation(
            bot, make_request(), pdf=b"%PDF-1.4", summary_in_group=True
        )
        bot.send_message.assert_not_awaited()
        bot.send_document.assert_not_awaited()

    async def test_admin_is_alerted_when_the_card_reached_nobody(self, tmp_paths, monkeypatch):
        """Заявка записана, карточки нет — тишины быть не должно.

        Именно так потерялась заявка при двухминутном провале WARP: сбой
        видел только лог.
        """
        from services import alerts

        monkeypatch.setattr(intake, "effective_finance_recipients", lambda: ["7"])
        seen = {}

        async def fake_alert(bot, title, details="", *, signature=None, kind=None, **_kw):
            seen["title"], seen["details"], seen["kind"] = title, details, kind
            return 1

        monkeypatch.setattr(alerts, "alert_admins", fake_alert)
        bot = self._bot()
        monkeypatch.setattr(intake, "notify_finance", AsyncMock(return_value=0))
        await intake.finalize_submission(bot, make_request(), invoice_file=None)
        assert "не дошла" in seen["title"]
        assert seen["kind"] == "delivery", "категория нужна: без неё алерт не выключить"

    async def test_admin_is_alerted_when_no_financiers_configured(self, tmp_paths, monkeypatch):
        from services import alerts

        monkeypatch.setattr(intake, "effective_finance_recipients", lambda: [])
        seen = {}

        async def fake_alert(bot, title, details="", *, signature=None, kind=None, **_kw):
            seen["title"], seen["kind"] = title, kind
            return 1

        monkeypatch.setattr(alerts, "alert_admins", fake_alert)
        monkeypatch.setattr(intake, "notify_finance", AsyncMock(return_value=0))
        await intake.finalize_submission(self._bot(), make_request(), invoice_file=None)
        assert "не настроены" in seen["title"] and seen["kind"] == "delivery"

    async def test_author_is_not_told_about_the_financier_failure(self):
        """Осечку чинит админ — сотруднику про неё не пишем."""
        bot = self._bot()
        await intake._send_user_confirmation(
            bot, make_request(urgency=Urgency.URGENT), pdf=b"%PDF-1.4", notified=0
        )
        caption = bot.send_document.await_args.kwargs["caption"]
        assert "не удалось" not in caption
        assert "администратору" not in caption

    async def test_author_still_hears_about_a_file_warning(self):
        """В групповую сводку предупреждения не пишут — автор узнаёт лично."""
        bot = self._bot()
        await intake._send_user_confirmation(
            bot,
            make_request(),
            pdf=b"%PDF-1.4",
            file_warning="⚠️ Сумма в счёте не совпала",
            summary_in_group=True,
        )
        bot.send_document.assert_not_awaited()
        assert "не совпала" in bot.send_message.await_args.kwargs["text"]

    async def test_without_a_group_the_author_gets_the_full_confirmation(self):
        bot = self._bot()
        await intake._send_user_confirmation(bot, make_request(), pdf=b"%PDF-1.4")
        bot.send_document.assert_awaited()
        assert "Заявка принята" in bot.send_document.await_args.kwargs["caption"]
