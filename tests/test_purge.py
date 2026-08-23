"""Очистка данных перед прод-запуском: отчёт считает, --apply стирает.

Проверяем главное свойство инструмента: без --apply он ничего не удаляет,
а с --apply не задевает то, что не выбрано (настройки бота, свежий
страховочный бэкап).
"""
from __future__ import annotations

import json

import pytest

from config import settings
from scripts import purge_data as purge
from services import cards, dedup, registry_sqlite, registry_xlsx, request_meta
from tests.conftest import make_request


@pytest.fixture()
def seeded(tmp_paths):
    """Каталог с документами, заявкой в реестре и её следами."""
    purge.ERRORS.clear()
    purge.SAFETY_BACKUP = None

    storage = settings.storage_path
    storage.mkdir(parents=True, exist_ok=True)
    (storage / "20260801_ООО Ромашка_1000.pdf").write_bytes(b"x" * 2048)
    (storage / "INV-20260801-120000-0001.pdf").write_bytes(b"z" * 1024)

    request = make_request()
    registry_sqlite.append_sync(request)
    registry_xlsx.append_sync(request, settings.registry_path)
    dedup.remember_sync(dedup.fingerprint(request), request.request_id)
    cards.save_sync(
        request_id=request.request_id,
        chat_id=-100,
        message_id=7,
        is_caption=False,
        base_html="карточка",
    )
    request_meta.save_reason_sync(request.request_id, "Отклонена", "нет счёта", "@fin")
    yield tmp_paths
    purge.ERRORS.clear()
    purge.SAFETY_BACKUP = None


def _scan(keys: str) -> dict[str, purge.Scan]:
    targets, unknown = purge.select_targets("documents", keys, "")
    assert unknown == []
    return {t.key: t.scan() for t in targets}


def test_scan_counts_documents_and_traces(seeded):
    scans = _scan("files,requests,dedup,cards,reasons,registry")

    assert scans["files"].count == 2
    assert scans["requests"].count == 1
    assert scans["dedup"].count == 1
    assert scans["cards"].count == 1
    assert scans["reasons"].count == 1
    assert scans["registry"].count == 1
    # Файлы не привязаны к заявке (file_url пустой) — это видно в отчёте.
    assert any("не привязано" in line for line in scans["files"].details)


def test_dry_run_deletes_nothing(seeded, capsys):
    targets, _ = purge.select_targets("documents", "", "")
    scans = {t.key: t.scan() for t in targets}
    purge.print_report(targets, scans, limit=5, full=False)

    out = capsys.readouterr().out
    assert "ничего не удалено" in out
    assert "Итого к удалению" in out
    assert list(settings.storage_path.iterdir())
    assert registry_sqlite.recent_requests_sync(10)


def test_apply_wipes_documents_and_keeps_settings(seeded):
    settings.runtime_settings_path.write_text(
        json.dumps({"finance": ["123"], "incidents": [{"kind": "x", "title": "y", "ts": 1.0}]}),
        encoding="utf-8",
    )

    targets, _ = purge.select_targets("all", "", "backups,users,conversations")
    result = purge.apply_purge(targets, quiet=True, no_backup=True, total=1)

    assert result["done"] is True
    assert not purge.ERRORS
    assert list(settings.storage_path.glob("*.pdf")) == []
    assert not settings.registry_path.exists()
    assert registry_sqlite.recent_requests_sync(10) == []
    assert cards.for_request_sync(make_request().request_id) == []
    assert dedup.check_sync(dedup.fingerprint(make_request())) is None

    # Состав финансистов — не документ: остаётся, журнал инцидентов чистится.
    saved = json.loads(settings.runtime_settings_path.read_text(encoding="utf-8"))
    assert saved["finance"] == ["123"]
    assert saved["incidents"] == []


def test_registry_recreated_after_wipe(seeded):
    targets, _ = purge.select_targets("documents", "registry", "")
    purge.apply_purge(targets, quiet=True, no_backup=True, total=1)
    assert not settings.registry_path.exists()

    registry_xlsx.append_sync(make_request(), settings.registry_path)

    from openpyxl import load_workbook

    ws = load_workbook(settings.registry_path).active
    assert ws.cell(row=1, column=1).value == "Дата внесения в реестр"
    assert ws.max_row == 2


def test_safety_backup_survives_backups_target(seeded):
    from services import backup

    old = backup.backup_dir() / "invoice-bot-backup-20260101-000000.tar.gz"
    old.parent.mkdir(parents=True, exist_ok=True)
    old.write_bytes(b"old")

    targets, _ = purge.select_targets("documents", "files,backups", "")
    result = purge.apply_purge(targets, quiet=True, no_backup=False, total=1)

    assert result["done"] is True
    assert not old.exists()
    assert purge.SAFETY_BACKUP is not None
    assert purge.SAFETY_BACKUP.exists()


def test_unknown_target_reported():
    targets, unknown = purge.select_targets("documents", "files,вчерашние", "")
    assert unknown == ["вчерашние"]
    assert [t.key for t in targets] == ["files"]


def test_skip_wins_over_preset():
    targets, unknown = purge.select_targets("documents", "", "files,registry")
    assert unknown == []
    assert "files" not in [t.key for t in targets]
    assert "requests" in [t.key for t in targets]
